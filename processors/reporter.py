"""
数据导出 v2 - 严格按表单表头对齐
表头: 店铺名称 | 平台 | 用户名 | 评分 | 评论内容 | 翻译内容 | 发布日期 | 采集时间 | 图片URLs | 商家回复 | 子评分 | 页面URL
"""
from __future__ import annotations
import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Optional

from core.models import Review

EXPORT_DIR = Path(__file__).parent.parent / "exports"

# ── 表单对齐的列定义 ──────────────────────────────────────────────
EXPORT_COLUMNS = [
    ("店铺名称",  lambda r: r.shop_name),
    ("平台",      lambda r: r.platform.value),
    ("用户名",    lambda r: r.reviewer_name),
    ("评分",      lambda r: r.rating),
    ("评论内容",  lambda r: r.content),
    ("翻译内容",  lambda r: r.translated_content or ""),
    ("发布日期",  lambda r: r.published_at.strftime("%Y-%m-%d %H:%M:%S") if r.published_at else ""),
    ("采集时间",  lambda r: r.crawled_at.strftime("%Y-%m-%d %H:%M:%S") if r.crawled_at else ""),
    ("图片URLs",  lambda r: "\n".join(r.image_urls) if r.image_urls else ""),
    ("商家回复",  lambda r: r.merchant_reply or ""),
    ("回复翻译",  lambda r: r.reply_translation or ""),
    ("子评分",    lambda r: _format_child_rating(r.child_rating)),
    ("页面URL",   lambda r: r.page_url or ""),
    # 扩展字段
    ("情感",      lambda r: r.sentiment.value if r.sentiment else ""),
    ("关键词",    lambda r: "、".join(r.keywords)),
    ("建议回复",  lambda r: r.suggested_reply or ""),
    ("爬取策略",  lambda r: r.ocr_strategy or ""),
]

HEADERS = [col[0] for col in EXPORT_COLUMNS]


def _format_child_rating(raw: Optional[str]) -> str:
    """将子评分 JSON 格式化为可读字符串"""
    if not raw:
        return ""
    try:
        d = json.loads(raw)
        return " | ".join(f"{k}:{v}" for k, v in d.items())
    except (json.JSONDecodeError, TypeError):
        return raw


def _row_data(r: Review) -> list:
    return [fn(r) for _, fn in EXPORT_COLUMNS]


def _ensure_dir() -> None:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ── CSV 导出 ──────────────────────────────────────────────────────
def export_csv(reviews: list[Review], filename: Optional[str] = None) -> Path:
    _ensure_dir()
    path = EXPORT_DIR / (filename or f"reviews_{_ts()}.csv")
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for r in reviews:
            writer.writerow(_row_data(r))
    return path


# ── JSON 导出 ─────────────────────────────────────────────────────
def export_json(reviews: list[Review], filename: Optional[str] = None) -> Path:
    _ensure_dir()
    path = EXPORT_DIR / (filename or f"reviews_{_ts()}.json")
    data = [dict(zip(HEADERS, _row_data(r))) for r in reviews]
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


# ── Excel 导出 ───────────────────────────────────────────────────
def export_excel(reviews: list[Review], filename: Optional[str] = None) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return export_csv(reviews, filename and filename.replace(".xlsx", ".csv"))

    _ensure_dir()
    path = EXPORT_DIR / (filename or f"reviews_{_ts()}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "评论数据"

    # 表头样式
    hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border

    ws.row_dimensions[1].height = 24

    # 情感颜色
    sentiment_fills = {
        "正面": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "负面": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "中性": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
    }
    content_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, r in enumerate(reviews, 2):
        row_data = _row_data(r)
        sentiment_val = r.sentiment.value if r.sentiment else ""
        row_fill = sentiment_fills.get(sentiment_val)

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = content_align
            cell.border = border
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 60

    # 列宽设置（按内容类型优化）
    col_widths = {
        "店铺名称": 20, "平台": 12, "用户名": 14, "评分": 8,
        "评论内容": 45, "翻译内容": 45, "发布日期": 18, "采集时间": 18,
        "图片URLs": 30, "商家回复": 40, "回复翻译": 40, "子评分": 25,
        "页面URL": 35, "情感": 8, "关键词": 25, "建议回复": 35, "爬取策略": 12,
    }
    for col_idx, header in enumerate(HEADERS, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = col_widths.get(header, 15)

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(HEADERS))}1"

    wb.save(path)
    return path
