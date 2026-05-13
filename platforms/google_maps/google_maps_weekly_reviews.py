from __future__ import annotations

import argparse
import asyncio
import csv
from datetime import datetime, timedelta
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
PROFILE_DIR = DATA / "browser_profiles" / "google_maps_public"
EXPORTS = ROOT / "exports" / "google_maps"
DEFAULT_EXCEL = Path(r"C:\Users\Administrator\Desktop\海外店鋪.xlsx")

FIELDS = [
    "Platform",
    "Country",
    "City",
    "JDE",
    "Store",
    "Google Rating",
    "Google Maps URL",
    "Resolved URL",
    "Review ID",
    "Reviewer",
    "Reviewer Local Guide",
    "Reviewer Review Count",
    "Rating",
    "Review Time Text",
    "Approx Review Date",
    "Within Last 7 Days",
    "Review Content",
    "Translated Snippet",
    "Translation Action Text",
    "Image URLs",
    "Owner Response",
    "Raw Text",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Google Maps weekly review collector for overseas HEYTEA stores")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL), help="Excel file with store rows")
    parser.add_argument("--sheet", default="各门店全渠道周报", help="Sheet name")
    parser.add_argument("--registry", default="", help="Unified store registry JSON; overrides Excel when provided")
    parser.add_argument("--country", default="", help="Only crawl country/city text containing this value")
    parser.add_argument("--shop-name", default="", help="Only crawl stores containing this value")
    parser.add_argument("--limit-stores", type=int, default=0, help="Limit stores; 0 means all")
    parser.add_argument("--days", type=int, default=7, help="Keep reviews within last N days")
    parser.add_argument("--max-scrolls", type=int, default=20, help="Review panel scroll attempts per store")
    parser.add_argument("--max-reviews-per-store", type=int, default=80, help="Maximum reviews parsed per store")
    parser.add_argument("--headless", action="store_true", help="Run browser headless")
    parser.add_argument("--browser-channel", default="msedge", help="Playwright browser channel")
    parser.add_argument("--output-prefix", default="google_maps_weekly_reviews", help="Export filename prefix")
    return parser.parse_args()


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def load_store_rows(args: argparse.Namespace) -> list[dict[str, Any]]:
    if args.registry:
        return load_store_rows_from_registry(args)
    path = Path(args.excel)
    if not path.exists():
        raise FileNotFoundError(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows: list[dict[str, Any]] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        city = clean_text(row[2] if len(row) > 2 else "")
        jde = clean_text(row[3] if len(row) > 3 else "")
        store = clean_text(row[5] if len(row) > 5 else "")
        rating = clean_text(row[10] if len(row) > 10 else "")
        url = clean_text(row[11] if len(row) > 11 else "")
        if not store or not jde or "google." not in url.lower():
            continue
        country = normalize_country(city, store)
        if args.country and args.country.lower() not in f"{country} {city}".lower():
            continue
        if args.shop_name and args.shop_name.lower() not in store.lower():
            continue
        rows.append(
            {
                "row_num": row_num,
                "country": country,
                "city": city,
                "jde": jde,
                "store": store,
                "google_rating": rating,
                "url": url,
            }
        )
    return rows


def load_store_rows_from_registry(args: argparse.Namespace) -> list[dict[str, Any]]:
    path = Path(args.registry)
    if not path.exists():
        raise FileNotFoundError(path)
    registry = json.loads(path.read_text(encoding="utf-8"))
    rows: list[dict[str, Any]] = []
    for index, store in enumerate(registry.get("stores", []), start=1):
        platform_data = (store.get("platforms") or {}).get("google_maps") or {}
        url = clean_text(platform_data.get("url", ""))
        if not url:
            continue
        country = clean_text(store.get("country", ""))
        city = clean_text(store.get("city", ""))
        store_name = clean_text(store.get("store_name", ""))
        jde = clean_text(store.get("jde", ""))
        if args.country and args.country.lower() not in f"{country} {city}".lower():
            continue
        if args.shop_name and args.shop_name.lower() not in f"{store_name} {jde}".lower():
            continue
        rows.append(
            {
                "row_num": store.get("source_row", index),
                "country": country,
                "city": city,
                "jde": jde,
                "store": store_name,
                "google_rating": clean_text(platform_data.get("meta", "")),
                "url": url,
            }
        )
    return rows


def normalize_country(city: str, store: str) -> str:
    text = f"{city} {store}"
    mapping = [
        ("澳门", ("澳门", "澳門")),
        ("香港", ("香港",)),
        ("加拿大", ("温哥华", "多伦多", "加拿大", "Burnaby", "Richmond")),
        ("马来西亚", ("吉隆坡", "雪兰莪", "槟城", "柔佛", "彭亨", "马来西亚")),
        ("美国", ("美国", "纽约", "洛杉矶", "旧金山", "波士顿", "西雅图", "圣地亚哥", "圣何塞", "湾区", "新泽西", "德州", "尔湾", "Virginia")),
        ("英国", ("英国", "伦敦", "曼彻斯特", "爱丁堡", "利兹", "南安普敦", "伯明翰", "利物浦")),
        ("韩国", ("韩国", "首尔", "大阪")),
        ("澳大利亚", ("澳大利亚", "墨尔本", "悉尼", "布里斯班")),
        ("新加坡", ("新加坡", "Vivo", "Jewel")),
    ]
    for country, tokens in mapping:
        if any(token in text for token in tokens):
            return country
    return city


def parse_relative_time(text: str, now: datetime, days: int) -> tuple[datetime | None, bool]:
    raw = clean_text(text).lower()
    if not raw:
        return None, False
    if any(token in raw for token in ("just now", "刚刚", "剛剛")):
        return now, True
    units = [
        (r"(\d+|a|an|one)\s*(minute|minutes|min)", "minutes"),
        (r"(\d+|a|an|one)\s*(hour|hours|hr)", "hours"),
        (r"(\d+|a|an|one)\s*(day|days)", "days"),
        (r"(\d+|a|an|one)\s*(week|weeks)", "weeks"),
        (r"(\d+|a|an|one)\s*(month|months)", "months"),
        (r"(\d+|a|an|one)\s*(year|years)", "years"),
        (r"(\d+)\s*(分钟|分鐘)", "minutes"),
        (r"(\d+)\s*(小时|小時)", "hours"),
        (r"(\d+)\s*天", "days"),
        (r"(\d+)\s*(周|週|星期|礼拜|禮拜)", "weeks"),
        (r"(\d+)\s*(个月|個月|月)", "months"),
        (r"(\d+)\s*年", "years"),
    ]
    for pattern, unit in units:
        match = re.search(pattern, raw)
        if not match:
            continue
        token = match.group(1)
        value = 1 if token in {"a", "an", "one"} else int(token)
        delta_days = {
            "minutes": value / 1440,
            "hours": value / 24,
            "days": value,
            "weeks": value * 7,
            "months": value * 30,
            "years": value * 365,
        }[unit]
        approx = now - timedelta(days=delta_days)
        return approx, delta_days <= days
    return None, False


async def click_first(page, selectors: tuple[str, ...], timeout: int = 1200) -> bool:
    for selector in selectors:
        try:
            locator = page.locator(selector).first
            if await locator.count() and await locator.is_visible(timeout=timeout):
                await locator.click(force=True)
                await page.wait_for_timeout(900)
                return True
        except Exception:
            pass
    return False


async def open_reviews_and_sort(page) -> None:
    await page.evaluate(
        """
        () => {
          const tabs = [...document.querySelectorAll('[role="tab"], button')];
          const target = tabs.find(el => {
            const text = (el.innerText || el.getAttribute('aria-label') || '').trim();
            if (/write|撰寫|撰写|我要/.test(text)) return false;
            return /Reviews|reviews|評論|評價|评价|评论/.test(text);
          });
          if (target) target.click();
        }
        """
    )
    await page.wait_for_timeout(1800)

    clicked_sort = await click_first(
        page,
        (
            "button[aria-label*='Sort reviews']",
            "button[aria-label*='Sort']",
            "button[aria-label*='排序']",
            "button:has-text('Sort')",
            "button:has-text('排序')",
        ),
        timeout=2500,
    )
    if clicked_sort:
        await click_first(
            page,
            (
                "div[role='menuitemradio']:has-text('Newest')",
                "div[role='menuitem']:has-text('Newest')",
                "div[role='menuitemradio']:has-text('最新')",
                "div[role='menuitem']:has-text('最新')",
                "text=Newest",
                "text=最新",
            ),
            timeout=2500,
        )
        await page.wait_for_timeout(1800)


async def expand_visible_reviews(page) -> None:
    """Expand visible review text and click read-only translation controls."""
    for _ in range(18):
        clicked = await page.evaluate(
            """
            () => {
              const expandButtons = [...document.querySelectorAll('button[jsaction*="expandReview"], button.w8nwRe')]
                .filter(btn => btn.offsetParent !== null);
              if (expandButtons.length) {
                expandButtons[0].click();
                return true;
              }
              const buttons = [...document.querySelectorAll('button, div[role="button"]')].filter(btn => {
                const text = (btn.innerText || btn.getAttribute('aria-label') || '').trim();
                if (!text || btn.offsetParent === null) return false;
                if (/撰写|撰寫|写评价|寫評價|Write a review|Reply|回复|回覆/i.test(text)) return false;
                return /^(More|more|展开|展開|更多)$/.test(text)
                  || /[.…]\\s*(More|more|更多)$/.test(text)
                  || /^(See translation|View translation|查看译文|查看譯文)/i.test(text)
                  || /^查看.*(译文|譯文)/.test(text);
              });
              const target = buttons[0];
              if (!target) return false;
              target.click();
              return true;
            }
            """
        )
        if not clicked:
            break
        await page.wait_for_timeout(250)


async def scroll_reviews(page, max_scrolls: int) -> None:
    for _ in range(max_scrolls):
        await expand_visible_reviews(page)
        before = await page.locator("[data-review-id], div.jftiEf").count()
        scrolled = await page.evaluate(
            """
            () => {
              const candidates = [...document.querySelectorAll('div[role="feed"], .m6QErb.DxyBCb, .m6QErb')];
              const el = candidates.find(x => x.scrollHeight > x.clientHeight + 100);
              if (!el) return false;
              el.scrollBy(0, Math.max(900, el.clientHeight * 0.9));
              return true;
            }
            """
        )
        if not scrolled:
            await page.mouse.wheel(0, 1200)
        await page.wait_for_timeout(1000)
        after = await page.locator("[data-review-id], div.jftiEf").count()
        if after == before:
            await page.wait_for_timeout(500)


async def parse_review_cards(page, store: dict[str, Any], days: int) -> list[dict[str, Any]]:
    now = datetime.now()
    rows = await page.evaluate(
        """
        ({store}) => {
          const cards = [...document.querySelectorAll('[data-review-id], div.jftiEf')];
          const seen = new Set();
          return cards.map((card, index) => {
            const reviewId = card.getAttribute('data-review-id') || '';
            const key = reviewId || card.innerText;
            if (seen.has(key)) return null;
            seen.add(key);
            const reviewerRaw = (card.querySelector('.d4r55, .WNxzHc, [class*="fontHeadlineSmall"]')?.innerText || '').trim();
            const reviewer = reviewerRaw.split('\\n')[0].trim();
            const ratingLabel = card.querySelector('span[role="img"][aria-label*="star"], span[aria-label*="star"], span[aria-label*="星"]')?.getAttribute('aria-label') || '';
            const timeText = (card.querySelector('.rsqaWe, .DU9Pgb .rsqaWe')?.innerText || '').trim();
            const contentCandidates = [...card.querySelectorAll('.wiI7pd, .MyEned span, .MyEned')]
              .map(el => (el.innerText || '').trim())
              .filter(Boolean)
              .map(text => text.replace(/[.…]\\s*(More|more|更多)$/g, '').trim())
              .filter(Boolean);
            const uniqueContents = [...new Set(contentCandidates)]
              .filter(text => !contentCandidates.some(other => other !== text && other.includes(text)));
            const content = (uniqueContents.sort((a, b) => b.length - a.length)[0] || '').trim();
            const translated = [...card.querySelectorAll('.kyuRq, .fontBodyMedium')]
              .map(el => (el.innerText || '').trim())
              .filter(text => /translated|translation|译文|譯文|翻译|翻譯|中文/i.test(text))
              .join('\\n')
              .trim();
            const translationAction = [...card.querySelectorAll('button, span, div[role="button"]')]
              .map(el => (el.innerText || el.getAttribute('aria-label') || '').trim())
              .filter(text => /translation|译文|譯文|翻译|翻譯|中文/i.test(text))
              .join('|');
            const localGuide = (card.innerText || '').includes('Local Guide') || (card.innerText || '').includes('本地向导') || (card.innerText || '').includes('在地嚮導');
            const reviewCountMatch = (card.innerText || '').match(/(\\d+[,.]?\\d*)\\s*(reviews|則評論|条评价|篇评价|則評價)/i);
            const imageUrls = [];
            card.querySelectorAll('button, a, div, img').forEach(el => {
              const style = el.getAttribute('style') || '';
              const src = el.currentSrc || el.src || '';
              const aria = el.getAttribute('aria-label') || '';
              const add = (url) => {
                if (!url || !/^https?:/.test(url)) return;
                if (!/googleusercontent|ggpht/.test(url)) return;
                if (/(=|-)w([1-8]?\\d)(-|$)|=s([1-8]?\\d)(-|$)|w36-h36|w40-h40|w48-h48|w60-h60|w64-h64|rp-mo|photo\\.jpg.*w=40/i.test(url)) return;
                if (!imageUrls.includes(url)) imageUrls.push(url);
              };
              if (src && !/profile|avatar/i.test(aria)) add(src);
              const match = style.match(/url\\(["']?([^"')]+)["']?\\)/);
              if (match) add(match[1]);
            });
            const owner = (card.querySelector('.CDe7pd, .wiI7pd + div')?.innerText || '').trim();
            return {
              Platform: 'Google Maps',
              Country: store.country || '',
              City: store.city || '',
              JDE: store.jde || '',
              Store: store.store || '',
              'Google Rating': store.google_rating || '',
              'Google Maps URL': store.url || '',
              'Resolved URL': location.href,
              'Review ID': reviewId,
              Reviewer: reviewer,
              'Reviewer Local Guide': localGuide ? 'YES' : '',
              'Reviewer Review Count': reviewCountMatch ? reviewCountMatch[1] : '',
              Rating: ratingLabel,
              'Review Time Text': timeText,
              'Review Content': content,
              'Translated Snippet': translated,
              'Translation Action Text': translationAction,
              'Image URLs': imageUrls.join('|'),
              'Owner Response': owner,
              'Raw Text': (card.innerText || '').trim(),
            };
          }).filter(Boolean);
        }
        """,
        {"store": store},
    )
    normalized: list[dict[str, Any]] = []
    for row in rows:
        approx, within = parse_relative_time(str(row.get("Review Time Text") or ""), now, days)
        row["Approx Review Date"] = approx.strftime("%Y-%m-%d %H:%M:%S") if approx else ""
        row["Within Last 7 Days"] = "YES" if within else ""
        normalized.append(row)
    return normalized


def dedupe(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    out: list[dict[str, Any]] = []
    for row in rows:
        key = row.get("Review ID") or f"{row.get('JDE')}:{row.get('Reviewer')}:{row.get('Review Time Text')}:{row.get('Review Content')[:80]}"
        if key in seen:
            continue
        seen.add(str(key))
        out.append(row)
    return out


async def crawl_store(page, store: dict[str, Any], args: argparse.Namespace) -> list[dict[str, Any]]:
    print(f"[store] {store['jde']} {store['store']}")
    await page.goto(store["url"], wait_until="domcontentloaded", timeout=60_000)
    await page.wait_for_timeout(4500)
    await open_reviews_and_sort(page)
    await scroll_reviews(page, args.max_scrolls)
    parsed = dedupe(await parse_review_cards(page, store, args.days))
    weekly = [row for row in parsed if row.get("Within Last 7 Days") == "YES"]
    rows = weekly[: args.max_reviews_per_store]
    print(f"[store] parsed_total={len(parsed)} weekly={len(rows)}")
    return rows


def export_rows(rows: list[dict[str, Any]], stores: list[dict[str, Any]], args: argparse.Namespace) -> tuple[Path, Path]:
    EXPORTS.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = EXPORTS / f"{args.output_prefix}_{stamp}.json"
    csv_path = EXPORTS / f"{args.output_prefix}_{stamp}.csv"
    payload = {
        "platform": "Google Maps",
        "source_excel": args.excel,
        "days": args.days,
        "store_count": len(stores),
        "review_count": len(rows),
        "stores": stores,
        "reviews": rows,
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return json_path, csv_path


async def run(args: argparse.Namespace) -> tuple[Path, Path, int]:
    from playwright.async_api import async_playwright

    stores = load_store_rows(args)
    if args.limit_stores:
        stores = stores[: args.limit_stores]
    print(f"[stores] {len(stores)}")

    async with async_playwright() as p:
        async def open_context():
            PROFILE_DIR.mkdir(parents=True, exist_ok=True)
            launch_options: dict[str, Any] = {
                "headless": args.headless,
                "viewport": {"width": 1280, "height": 920},
                "locale": "zh-HK",
                "timezone_id": "Asia/Hong_Kong",
            }
            if args.browser_channel:
                launch_options["channel"] = args.browser_channel
            new_context = await p.chromium.launch_persistent_context(str(PROFILE_DIR), **launch_options)
            new_page = new_context.pages[0] if new_context.pages else await new_context.new_page()
            return new_context, new_page

        context, page = await open_context()
        all_rows: list[dict[str, Any]] = []
        try:
            for store in stores:
                for attempt in range(2):
                    try:
                        if page.is_closed():
                            page = await context.new_page()
                        all_rows.extend(await crawl_store(page, store, args))
                        break
                    except Exception as exc:
                        message = str(exc)
                        print(f"[store:error] {store.get('jde')} {store.get('store')}: {message}")
                        if attempt == 0 and "Target page, context or browser has been closed" in message:
                            try:
                                await context.close()
                            except Exception:
                                pass
                            context, page = await open_context()
                            continue
                        break
            json_path, csv_path = export_rows(all_rows, stores, args)
            return json_path, csv_path, len(all_rows)
        finally:
            try:
                await context.close()
            except Exception:
                pass


def main() -> None:
    args = parse_args()
    json_path, csv_path, count = asyncio.run(run(args))
    print(f"[done] reviews={count}")
    print(f"[done] json={json_path}")
    print(f"[done] csv={csv_path}")


if __name__ == "__main__":
    main()
