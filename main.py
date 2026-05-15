"""
外卖评论爬虫系统 - NiceGUI Web 主入口
运行: python main.py
访问: http://localhost:8080
"""
import base64
import json
import difflib
import os
import re
import sys
import threading
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from uuid import uuid4

# 确保项目根目录在 PYTHONPATH
sys.path.insert(0, str(Path(__file__).parent))

from dataclasses import replace
from fastapi import Request
from nicegui import ui, app as nicegui_app
from core.database import init_db
from ui.pages import dashboard, crawler, global_ops, reviews, reports, settings, sentiment
from unified_collector.coordinator import COORDINATOR
from unified_collector.monitor import EVENT_BUS, SYNC_MONITOR, validate_week_range
from unified_collector.platform_capabilities import PLATFORM_CAPABILITIES, canonical_platform
from unified_collector.settings import (
    chat_with_configured_provider,
    load_settings,
    reset_settings,
    save_settings,
    smoke_check_settings,
    smoke_test_provider,
)
from unified_collector.task_loader import load_task

ROOT = Path(__file__).parent
STITCH_STATIC_DIR = ROOT / "ui" / "stitch_static"
STORE_REGISTRY_PATH = ROOT / "data" / "store_registry.json"
TASK_DIR = ROOT / "unified_collector" / "tasks"
EXPORT_DIR = ROOT / "exports"
RUNS_DIR = EXPORT_DIR / "runs"
KNOWLEDGE_DIR = ROOT / "data" / "knowledge_base"
URL_PATTERN = re.compile(r"https?://[^\s\"'<>]+")
MONEY_PATTERN = re.compile(r"(?:[$¥￥€£₩]|HK\$|US\$|C\$|A\$|SG\$|RM)?\s*-?\d{1,6}(?:[,.]\d{1,2})?")
ORDER_ID_PATTERNS = (
    re.compile(r"(?:订单号|訂單號|order\s*(?:id|no\.?|number|#))\s*[:：#]?\s*([A-Z]{0,6}\d{6,})", re.IGNORECASE),
    re.compile(r"\b([A-Z]{1,6}\d{7,})\b"),
)
_TRANSLATION_CACHE: dict[str, str] = {}
_TRANSLATION_CACHE_LOCK = threading.RLock()
_TRANSLATION_CACHE_MAX = 2000

try:
    from opencc import OpenCC  # type: ignore

    _OPENCC_T2S = OpenCC("t2s")
except Exception:
    _OPENCC_T2S = None

_TRADITIONAL_HINT_CHARS = set(
    "體臺萬與為說評價後廣門點開關顧飲龍東線區應轉譯圖單訂這裝種級態務條數據難壞時過優舖們嗎頁輸綜評論"
    "對於沒麼讓將會個員請處層發現滿意推薦變溫遞遲遜舊嚴靜雲灣價錢裡還"
    "麼麼師灣灣與於對沒請讓應該實際驗證傳遞選擇發貨聯繫態度質量衛生"
)
_BASIC_T2S_CHAR_MAP = str.maketrans(
    {
        "體": "体",
        "臺": "台",
        "萬": "万",
        "與": "与",
        "為": "为",
        "說": "说",
        "評": "评",
        "價": "价",
        "後": "后",
        "廣": "广",
        "門": "门",
        "點": "点",
        "開": "开",
        "關": "关",
        "顧": "顾",
        "飲": "饮",
        "龍": "龙",
        "東": "东",
        "線": "线",
        "區": "区",
        "應": "应",
        "轉": "转",
        "譯": "译",
        "圖": "图",
        "單": "单",
        "訂": "订",
        "這": "这",
        "裝": "装",
        "種": "种",
        "級": "级",
        "態": "态",
        "務": "务",
        "條": "条",
        "數": "数",
        "據": "据",
        "難": "难",
        "壞": "坏",
        "時": "时",
        "過": "过",
        "優": "优",
        "舖": "铺",
        "們": "们",
        "頁": "页",
        "輸": "输",
        "綜": "综",
        "對": "对",
        "於": "于",
        "沒": "没",
        "麼": "么",
        "讓": "让",
        "將": "将",
        "會": "会",
        "個": "个",
        "員": "员",
        "請": "请",
        "處": "处",
        "層": "层",
        "發": "发",
        "現": "现",
        "滿": "满",
        "溫": "温",
        "遞": "递",
        "遲": "迟",
        "遜": "逊",
        "嚴": "严",
        "靜": "静",
        "雲": "云",
        "錢": "钱",
        "裡": "里",
        "聯": "联",
        "繫": "系",
        "衛": "卫",
        "質": "质",
        "選": "选",
        "擇": "择",
        "傳": "传",
        "實": "实",
        "驗": "验",
        "證": "证",
        "標": "标",
        "準": "准",
        "門": "门",
        "點": "点",
        "舖": "铺",
        "舊": "旧",
        "還": "还",
    }
)

# ── 初始化数据库 ──────────────────────────────────────────────────
init_db()

if STITCH_STATIC_DIR.exists():
    nicegui_app.add_static_files("/stitch-static", STITCH_STATIC_DIR, max_cache_age=0)


@nicegui_app.middleware("http")
async def stitch_static_no_cache(request: Request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/stitch-static/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


def _safe_platform_data(data: dict) -> dict:
    """Expose only non-secret platform metadata to the UI."""
    safe: dict[str, object] = {}
    if data.get("url"):
        safe["url"] = data["url"]
    if data.get("meta"):
        safe["meta"] = data["meta"]
    if data.get("store_id"):
        safe["store_id"] = data["store_id"]
    if data.get("note"):
        note = str(data["note"])
        lowered = note.lower()
        if not any(token in lowered for token in ("密码", "password", "pwd", "账号", "account", "token", "key")):
            safe["note"] = note
        else:
            safe["note"] = "[local credential note hidden]"
    if data.get("account_ref"):
        safe["account_ref"] = data["account_ref"]
    return safe


def _load_safe_registry() -> dict:
    if not STORE_REGISTRY_PATH.exists():
        return {"schema_version": 1, "store_count": 0, "stores": [], "platform_counts": {}}
    raw = json.loads(STORE_REGISTRY_PATH.read_text(encoding="utf-8"))
    stores = []
    platform_counts: dict[str, int] = {}
    for store in raw.get("stores", []):
        platforms = {}
        for name, pdata in (store.get("platforms") or {}).items():
            safe_data = _safe_platform_data(pdata or {})
            if safe_data:
                platforms[name] = safe_data
                platform_counts[name] = platform_counts.get(name, 0) + 1
        stores.append(
            {
                "jde": store.get("jde", ""),
                "store_name": store.get("store_name", ""),
                "country": _normalize_region_name(str(store.get("country", ""))),
                "country_code": store.get("country_code", ""),
                "city": store.get("city", ""),
                "platforms": platforms,
            }
        )
    return {
        "schema_version": raw.get("schema_version", 1),
        "store_count": len(stores),
        "stores": stores,
        "platform_counts": dict(sorted(platform_counts.items())),
        "generated_at": raw.get("generated_at", ""),
    }


def _task_files() -> list[Path]:
    return sorted(path for path in TASK_DIR.glob("*.json") if path.is_file()) if TASK_DIR.exists() else []


def _as_list(value: Any) -> list[Any]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                return parsed
        except Exception:
            pass
        return [part.strip() for part in value.replace("\n", "|").split("|") if part.strip()]
    return [value]


def _is_mostly_chinese(text: str) -> bool:
    value = str(text or "").strip()
    if not value:
        return False
    chinese = sum(1 for ch in value if "\u4e00" <= ch <= "\u9fff")
    return chinese / max(1, len(value)) >= 0.3


def _looks_traditional_chinese(text: str) -> bool:
    value = str(text or "")
    return any(ch in _TRADITIONAL_HINT_CHARS for ch in value)


def _to_simplified_chinese_local(text: str) -> str:
    value = str(text or "")
    if not value or not _is_mostly_chinese(value):
        return value
    converted = value
    if _OPENCC_T2S is not None:
        try:
            converted = _OPENCC_T2S.convert(converted)
        except Exception:
            converted = value
    converted = converted.translate(_BASIC_T2S_CHAR_MAP)
    return converted


async def _to_simplified_chinese(text: str, force_remote: bool = False) -> str:
    value = str(text or "")
    if not value:
        return value
    local = _to_simplified_chinese_local(value)
    if not force_remote and (local != value or not _looks_traditional_chinese(value)):
        return local
    try:
        prompt = (
            "请将下面文本转换为简体中文，只输出转换后的文本，不要解释：\n\n"
            f"{value[:5000]}"
        )
        remote = (await chat_with_configured_provider(prompt, max_tokens=1200)).strip()
        if remote:
            remote = _to_simplified_chinese_local(remote)
        return remote or local
    except Exception:
        return local


def _extract_nested_order_fields(raw: dict[str, Any]) -> tuple[list[dict[str, Any]], str]:
    item_containers: list[Any] = []
    detail_texts: list[str] = []
    seen_refs: set[int] = set()
    item_key_tokens = ("item", "product", "sku", "goods", "dish", "menu", "line")
    detail_key_tokens = (
        "orderdetail",
        "order_detail",
        "orderdetails",
        "orderitems",
        "ordereditems",
        "productview",
        "items",
        "products",
        "goods",
        "remark",
    )
    excluded_key_tokens = (
        "reviewitemdetail",
        "reviewdetail",
        "subrating",
        "subratings",
        "ratingdetail",
        "scoreitem",
        "qualityflag",
        "reviewtype",
        "reviewquality",
        "sourcefile",
    )

    def walk(node: Any, depth: int = 0) -> None:
        if depth > 6:
            return
        ref_id = id(node)
        if ref_id in seen_refs:
            return
        seen_refs.add(ref_id)
        if isinstance(node, dict):
            for key, value in node.items():
                key_l = str(key or "").lower().replace(" ", "").replace("-", "").replace("_", "")
                if any(token in key_l for token in excluded_key_tokens):
                    continue
                if any(token in key_l for token in detail_key_tokens) and not isinstance(value, str):
                    item_containers.append(value)
                if isinstance(value, (dict, list, tuple)):
                    walk(value, depth + 1)
                elif isinstance(value, str):
                    text = value.strip()
                    excluded_context = any(
                        token in key_l
                        for token in (
                            "reviewer",
                            "customer",
                            "author",
                            "avatar",
                            "area",
                            "country",
                            "restaurant",
                            "store",
                            "branch",
                            "platform",
                        )
                    )
                    order_context = any(
                        token in key_l
                        for token in (
                            "order",
                            "item",
                            "product",
                            "goods",
                            "sku",
                            "dish",
                            "menu",
                            "qty",
                            "quantity",
                            "count",
                            "price",
                            "amount",
                            "subtotal",
                            "remark",
                        )
                    )
                    if text and order_context and not excluded_context:
                        detail_texts.append(text)
        elif isinstance(node, (list, tuple)):
            if node and all(isinstance(item, dict) for item in node):
                keys_flat = " ".join(" ".join(str(key).lower() for key in item.keys()) for item in node[:5])
                if any(token in keys_flat for token in item_key_tokens):
                    item_containers.append(list(node))
            for value in node:
                walk(value, depth + 1)

    walk(raw)
    extracted_items: list[dict[str, Any]] = []
    for candidate in item_containers:
        parsed = _parse_ordered_items(candidate)
        for item in parsed:
            if item not in extracted_items:
                extracted_items.append(item)
    merged_detail = "\n".join(text for text in detail_texts if text)[:12000]
    return extracted_items, merged_detail


def _extract_urls_from_value(value: Any) -> list[str]:
    urls: list[str] = []
    for item in _as_list(value):
        if isinstance(item, dict):
            for sub in item.values():
                urls.extend(_extract_urls_from_value(sub))
            continue
        text = str(item or "")
        if not text:
            continue
        if "|" in text:
            parts = [part.strip() for part in text.split("|") if part.strip()]
            for part in parts:
                urls.extend(_extract_urls_from_value(part))
            continue
        urls.extend(URL_PATTERN.findall(text))
    # stable de-dup
    seen: set[str] = set()
    deduped: list[str] = []
    for url in urls:
        if url in seen:
            continue
        seen.add(url)
        deduped.append(url)
    return deduped


def _collect_image_urls(raw: dict[str, Any]) -> list[str]:
    urls: list[str] = []
    fields = (
        "image_urls",
        "Image URLs",
        "Review Image URLs",
        "Photo URLs",
        "photos",
        "Product Image URLs",
        "Product Images",
        "商品图",
    )
    for field in fields:
        if field in raw and raw.get(field) not in (None, ""):
            urls.extend(_extract_urls_from_value(raw.get(field)))
    # fallback: parse from order detail text
    if not urls:
        for field in ("order_detail", "Order Detail", "Order Details", "Expanded Order Detail"):
            text = raw.get(field)
            if text:
                urls.extend(_extract_urls_from_value(text))
    seen: set[str] = set()
    deduped: list[str] = []
    for url in urls:
        if url in seen:
            continue
        # drop avatar-like tiny profile images when key words are explicit
        lowered = url.lower()
        if "avatar" in lowered or "profile" in lowered:
            continue
        seen.add(url)
        deduped.append(url)
    return deduped


def _first(raw: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        value = raw.get(key)
        if value not in (None, "", []):
            return value
    return default


def _parse_ordered_items(value: Any) -> list[dict[str, Any]]:
    def _is_subrating_text(text: str) -> bool:
        raw = str(text or "").strip()
        if not raw:
            return False
        has_rating_labels = bool(re.search(r"(綜合|综合|口味|包裝|包装|配送|服務|服务)\s*[:：]\s*\d", raw))
        has_item_signal = bool(
            re.search(
                r"([xX×]\s*\d+|商品|產品|产品|菜品|饮品|飲品|order\s*item|menu\s*item|product|item|goods)",
                raw,
                flags=re.IGNORECASE,
            )
        )
        return has_rating_labels and not has_item_signal

    def _extract_money_token(text: str) -> str:
        raw = str(text or "").strip()
        if not raw:
            return ""
        tokens = [match.group(0).strip() for match in MONEY_PATTERN.finditer(raw)]
        candidates = [
            token for token in tokens
            if token and any(ch.isdigit() for ch in token) and len(token.replace(",", "").replace(".", "").strip()) >= 1
        ]
        if not candidates:
            return ""
        for token in reversed(candidates):
            lowered = token.lower()
            if "x" in lowered and not any(symbol in token for symbol in ("$", "¥", "￥", "€", "£", "₩")):
                continue
            return token.replace(" ", "")
        return candidates[-1].replace(" ", "")

    def _parse_quantity(text: str) -> str:
        raw = str(text or "")
        if not raw:
            return ""
        for pattern in (r"[xX×]\s*(\d+)", r"数量[:：]?\s*(\d+)", r"qty[:：]?\s*(\d+)", r"count[:：]?\s*(\d+)"):
            match = re.search(pattern, raw, flags=re.IGNORECASE)
            if match:
                return match.group(1)
        return ""

    def _normalize_item(item: dict[str, Any]) -> dict[str, Any]:
        normalized = dict(item)
        text_candidates = [
            str(normalized.get(key) or "")
            for key in ("text", "name", "item", "itemName", "spec", "specs", "desc", "detail", "remark")
        ]
        if any(_is_subrating_text(text) for text in text_candidates):
            return {}
        price_keys = (
            "unit_price",
            "unitPrice",
            "price",
            "priceStr",
            "amount",
            "totalPrice",
            "productPrice",
            "goodsPrice",
            "salePrice",
            "finalPrice",
            "singlePrice",
            "subtotal",
            "order_total",
        )
        price = ""
        for key in price_keys:
            value = normalized.get(key)
            if value in (None, ""):
                continue
            value_text = str(value).strip()
            if value_text:
                price = value_text
                break
        if not price:
            for key in ("text", "name", "item", "itemName", "spec", "specs", "desc", "detail", "remark"):
                value = normalized.get(key)
                if value in (None, ""):
                    continue
                price = _extract_money_token(str(value))
                if price:
                    break
        if price and not normalized.get("unit_price"):
            normalized["unit_price"] = price

        qty = ""
        for key in ("quantity", "qty", "count", "num", "productCount"):
            value = normalized.get(key)
            if value in (None, ""):
                continue
            qty = str(value).strip()
            if qty:
                break
        if not qty:
            for key in ("text", "name", "item", "desc", "detail"):
                value = normalized.get(key)
                if value in (None, ""):
                    continue
                qty = _parse_quantity(str(value))
                if qty:
                    break
        if qty and not normalized.get("quantity"):
            normalized["quantity"] = qty
        return normalized

    def _parse_item_line(text: str) -> dict[str, Any]:
        line = str(text or "").strip()
        if not line:
            return {}
        if _is_subrating_text(line):
            return {}
        price = _extract_money_token(line)
        qty = _parse_quantity(line)
        has_item_marker = bool(
            re.search(
                r"(商品|產品|产品|菜品|饮品|飲品|订购清单|訂購清單|order\s*item|menu\s*item|product|item|goods)",
                line,
                flags=re.IGNORECASE,
            )
        )
        if not (price or qty or has_item_marker):
            return {}
        name = line
        if price:
            name = name.replace(price, " ")
        for token in (r"[xX×]\s*\d+", r"数量[:：]?\s*\d+", r"qty[:：]?\s*\d+", r"count[:：]?\s*\d+"):
            name = re.sub(token, " ", name, flags=re.IGNORECASE)
        name = re.sub(r"\s+", " ", name).strip(" -:：|")
        payload: dict[str, Any] = {"name": name or line, "text": line}
        if qty:
            payload["quantity"] = qty
        if price:
            payload["unit_price"] = price
        return payload

    parsed_value = value
    if isinstance(parsed_value, str):
        stripped = parsed_value.strip()
        if stripped.startswith("{") or stripped.startswith("["):
            try:
                parsed_value = json.loads(stripped)
            except Exception:
                parsed_value = value
        else:
            text_lines = [line.strip(" ;|") for line in re.split(r"[\r\n;]+", stripped) if line.strip(" ;|")]
            parsed_items = [_parse_item_line(line) for line in text_lines]
            parsed_items = [item for item in parsed_items if item]
            if parsed_items:
                return [_normalize_item(item) for item in parsed_items]
    if isinstance(parsed_value, dict):
        # Platform-specific order detail payloads often nest line items in `details`.
        for key in ("details", "items", "order_items", "orderItemList"):
            nested = parsed_value.get(key)
            if isinstance(nested, list):
                parsed_value = nested
                break
    items = _as_list(parsed_value)
    if len(items) == 1 and isinstance(items[0], str):
        stripped = items[0].strip()
        if stripped.startswith("{") or stripped.startswith("["):
            try:
                parsed = json.loads(stripped)
                if isinstance(parsed, list):
                    items = parsed
                elif isinstance(parsed, dict):
                    for key in ("details", "items", "order_items", "orderItemList"):
                        nested = parsed.get(key)
                        if isinstance(nested, list):
                            items = nested
                            break
            except Exception:
                pass
    normalized = []
    for item in items:
        if isinstance(item, dict):
            normalized_item = _normalize_item(item)
            if normalized_item:
                normalized.append(normalized_item)
        elif item not in (None, ""):
            parsed = _parse_item_line(str(item))
            normalized_item = _normalize_item(parsed or {"text": str(item)})
            if normalized_item:
                normalized.append(normalized_item)
    return normalized


def _extract_order_id_from_text(value: Any) -> str:
    text = str(value or "")
    if not text:
        return ""
    for pattern in ORDER_ID_PATTERNS:
        match = pattern.search(text)
        if match:
            return match.group(1).strip()
    return ""


def _extract_reviews_from_payload(payload: Any) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    if isinstance(payload, list):
        return {}, [item for item in payload if isinstance(item, dict)]
    if not isinstance(payload, dict):
        return {}, []
    for key in ("reviews", "data", "items", "rows", "records"):
        value = payload.get(key)
        if isinstance(value, list):
            return payload, [item for item in value if isinstance(item, dict)]
    return payload, []


def _normalize_ui_review(raw: dict[str, Any], payload: dict[str, Any], source_file: Path, index: int) -> dict[str, Any]:
    platform = _first(raw, "platform", "Platform", default=payload.get("platform", ""))
    country = _first(raw, "country", "Country", "region", "Region", default=payload.get("country", payload.get("region", payload.get("Region", ""))))
    store = _first(raw, "store", "Store", "shop", "Shop", "Branch", "branch", "store_name", "Store Name")
    rating = _first(raw, "rating", "Rating", "score")
    review = _first(
        raw,
        "review",
        "Review contents",
        "Review Content",
        "content",
        "Content",
        "Comment",
        "Review",
        "raw_text",
        "Raw Text",
        "评价内容",
        "评价",
    )
    translated = _first(
        raw,
        "translated_review",
        "Translated Review",
        "translated_content",
        "Google translation",
        "Chinese Translation",
        "CN Translation",
    )
    review_time = _first(raw, "review_time", "Review time", "Review Time", "Approx Review Date", "Review Date", "time", "Date", "timestamp")
    image_urls = _collect_image_urls(raw)
    ordered_items = _parse_ordered_items(
        _first(
            raw,
            "ordered_items",
            "Ordered Items",
            "Order Items JSON",
            "Order Items Text",
            "items",
            "Products",
            "Order Items",
            "Order Detail JSON",
        )
    )
    order_detail = _first(
        raw,
        "order_detail",
        "Order Detail",
        "Order Details",
        "Expanded Order Detail",
        "order_details",
        "Order Detail JSON",
    )
    nested_items, nested_detail = _extract_nested_order_fields(raw)
    if not ordered_items and nested_items:
        ordered_items = nested_items
    if (not order_detail or str(order_detail).strip() in {"", "-", "null", "None"}) and nested_detail:
        order_detail = nested_detail
    order_detail = _clean_order_detail_text(order_detail)
    if _is_noise_order_detail(order_detail):
        order_detail = ""
    order_id = _first(raw, "order_id", "Order ID", "Order View ID")
    order_total = _first(raw, "order_total", "Order Total", "Total", "Subtotal", "订单金额", "訂單金額")
    if not _meaningful(order_id):
        order_id = _extract_order_id_from_text(order_detail) or _extract_order_id_from_text(json.dumps(raw, ensure_ascii=False)[:6000])
    quality_flags = _as_list(_first(raw, "quality_flags", "Quality Flags"))
    review_id = _first(raw, "review_id", "id", "Review ID", default=f"{source_file.stem}-{index}")
    review_text = _valid_review_text(review)
    translated_text = _to_simplified_chinese_local(str(translated))
    if not _valid_review_text(translated_text):
        translated_text = ""
    if not translated_text and review_text and _is_mostly_chinese(review_text):
        translated_text = _to_simplified_chinese_local(review_text)
    return {
        "review_id": str(review_id),
        "run_id": _first(raw, "run_id", default=payload.get("run_id", "")),
        "platform": str(platform),
        "country": _normalize_region_name(str(country)),
        "store": str(store),
        "store_id": str(_first(raw, "store_id", "Store ID", "branch_id", "jde", "JDE")),
        "rating": rating,
        "sub_ratings": _first(raw, "sub_ratings", "Child ratings", "child_ratings", "Sub Ratings"),
        "review": review_text,
        "review_language": str(_first(raw, "review_language", "language", "Language")),
        "translated_review": translated_text,
        "customer": _clean_customer_display(_first(raw, "customer", "Customer", "reviewer", "Reviewer", "Reviewer Name", "user_name", "User Name")),
        "review_time": str(review_time),
        "order_id": str(order_id),
        "ordered_items": ordered_items,
        "order_detail": str(order_detail),
        "order_total": str(order_total),
        "image_urls": image_urls,
        "source": str(_first(raw, "source", "Source", default="jsonl" if source_file.suffix == ".jsonl" else "json")),
        "source_file": str(source_file.relative_to(ROOT)),
        "quality_flags": [str(flag) for flag in quality_flags],
        "raw_json": raw,
        "has_order": bool(order_id or order_detail or ordered_items),
        "has_image": bool(image_urls),
    }


def _review_sort_key(record: dict[str, Any]) -> str:
    return str(record.get("review_time") or "") + "|" + str(record.get("source_file") or "")


def _clean_identity_text(value: Any, max_len: int = 220) -> str:
    text = str(value or "").strip()
    if text in {"", "-", "None", "none", "null", "NULL", "No data", "no data", "N/A", "n/a"}:
        return ""
    text = _to_simplified_chinese_local(text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text[:max_len]


def _review_identity_key(record: dict[str, Any]) -> str:
    platform = canonical_platform(str(record.get("platform") or "")).lower()
    if not platform:
        platform = _clean_identity_text(record.get("platform"), 80)
    order_id = _clean_identity_text(record.get("order_id") or record.get("order_sn"), 120)
    if order_id:
        return f"order|{platform}|{order_id}"

    store = _clean_identity_text(record.get("store") or record.get("store_id"), 160)
    customer = _clean_identity_text(record.get("customer"), 120)
    rating = _clean_identity_text(record.get("rating"), 20)
    if platform == "google_maps" and store and customer:
        return f"google_user|{platform}|{store}|{customer}|{rating}"

    review_text = _clean_identity_text(record.get("review") or record.get("translated_review"), 260)
    if review_text:
        return f"text|{platform}|{store}|{customer}|{rating}|{review_text}"

    review_id = _clean_identity_text(record.get("review_id"), 160)
    source_file = _clean_identity_text(record.get("source_file"), 240)
    review_id_base = re.sub(r"[-_]\d{1,6}$", "", review_id)
    if review_id and source_file and review_id not in source_file and review_id_base not in source_file and len(review_id) >= 10:
        return f"review_id|{platform}|{review_id}"

    review_time = _clean_identity_text(str(record.get("review_time") or "")[:10], 40)
    if review_time or customer or store or rating:
        return f"rating_only|{platform}|{review_time}|{customer}|{rating}|{store}"
    return f"fallback|{platform}|{source_file}"


def _review_fuzzy_base_key(record: dict[str, Any]) -> str:
    platform = canonical_platform(str(record.get("platform") or "")).lower()
    if not platform:
        platform = _clean_identity_text(record.get("platform"), 80)
    store = _clean_identity_text(record.get("store") or record.get("store_id"), 160)
    rating = _clean_identity_text(record.get("rating"), 20)
    review_time = _clean_identity_text(str(record.get("review_time") or "")[:16], 40)
    return f"{platform}|{store}|{rating}|{review_time}"


def _review_text_identity(record: dict[str, Any], max_len: int = 2000) -> str:
    text = str(record.get("review") or record.get("translated_review") or "")
    text = re.sub(r"[\ue000-\uf8ff]", " ", text)
    text = re.sub(r"(?:…|\.\.\.)\s*(?:更多|more|show\s+more).*$", "", text, flags=re.IGNORECASE | re.DOTALL)
    return _clean_identity_text(text, max_len)


def _same_review_text(left: str, right: str) -> bool:
    if not left or not right:
        return False
    if left == right:
        return True
    shorter, longer = sorted((left, right), key=len)
    if len(shorter) < 40:
        return False
    if shorter in longer:
        return True
    if len(shorter) >= 80 and shorter[:96] == longer[:96]:
        return True
    return difflib.SequenceMatcher(None, shorter[:800], longer[:800]).ratio() >= 0.92


def _same_review_entity(left: dict[str, Any], right: dict[str, Any]) -> bool:
    if _review_fuzzy_base_key(left) != _review_fuzzy_base_key(right):
        return False
    left_customer = _clean_identity_text(left.get("customer"), 120)
    right_customer = _clean_identity_text(right.get("customer"), 120)
    if left_customer and right_customer and left_customer != right_customer:
        return False
    left_order = _clean_identity_text(left.get("order_id"), 160)
    right_order = _clean_identity_text(right.get("order_id"), 160)
    if left_order and right_order:
        return left_order == right_order
    left_text = _review_text_identity(left)
    right_text = _review_text_identity(right)
    if left_text or right_text:
        return _same_review_text(left_text, right_text)
    return True


def _resolve_review_dedupe_key(
    records_by_key: dict[str, dict[str, Any]],
    fuzzy_buckets: dict[str, list[str]],
    record: dict[str, Any],
) -> str:
    key = _review_identity_key(record)
    if key in records_by_key:
        return key
    for candidate_key in fuzzy_buckets.get(_review_fuzzy_base_key(record), []):
        candidate = records_by_key.get(candidate_key)
        if candidate and _same_review_entity(record, candidate):
            return candidate_key
    if not key.startswith("text|"):
        return key
    text = _review_text_identity(record)
    if not text:
        return key
    for candidate_key in fuzzy_buckets.get(_review_fuzzy_base_key(record), []):
        candidate = records_by_key.get(candidate_key)
        if candidate and _same_review_text(text, _review_text_identity(candidate)):
            return candidate_key
    return key


def _remember_review_dedupe_key(fuzzy_buckets: dict[str, list[str]], record: dict[str, Any], key: str) -> None:
    bucket = fuzzy_buckets.setdefault(_review_fuzzy_base_key(record), [])
    if key not in bucket:
        bucket.append(key)


def _meaningful(value: Any) -> bool:
    if value in (None, "", [], {}):
        return False
    text = str(value).strip()
    return text not in {"", "-", "None", "none", "null", "NULL", "No data", "no data", "N/A", "n/a"}


def _is_noise_order_detail(value: Any) -> bool:
    text = str(value or "").strip()
    if not _meaningful(text):
        return True
    text = _clean_order_detail_text(text)
    if not _meaningful(text):
        return True
    markers = (
        "order",
        "item",
        "qty",
        "quantity",
        "price",
        "subtotal",
        "total",
        "product",
        "goods",
        "订单",
        "商品",
        "数量",
        "单价",
        "价格",
        "小计",
        "结算",
    )
    lower = text.lower()
    has_marker = any(marker.lower() in lower for marker in markers)
    has_money = bool(re.search(r"(?:[$€£¥]|HK\$|MOP|RM|SGD|USD|AUD|CAD)\s*\d", text, flags=re.IGNORECASE))
    has_quantity = bool(re.search(r"(?:^|\s)(?:x\s*)?\d+\s*(?:份|杯|件|pcs?|items?)\b", text, flags=re.IGNORECASE))
    return not (has_marker or has_money or has_quantity)


def _clean_order_detail_text(value: Any) -> str:
    lines = []
    for line in str(value or "").replace("\r", "\n").split("\n"):
        text = line.strip()
        if not text:
            continue
        lower = text.lower()
        if re.fullmatch(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}", text):
            continue
        if ".owner" in lower or "owner a" in lower or "owner b" in lower:
            continue
        if re.fullmatch(r"[A-Za-z]:[\\/].+", text) or re.search(r"exports[\\/].+\.jsonl?$", text, flags=re.IGNORECASE):
            continue
        lines.append(text)
    return "\n".join(lines).strip()


def _valid_review_text(value: Any) -> str:
    text = str(value or "").strip()
    if not _meaningful(text):
        return ""
    if re.fullmatch(r"[\d\s:：\-.,，。/]+", text):
        return ""
    return text


def _clean_customer_display(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\b[a-fA-F0-9]{16,64}\b", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _review_completeness_score(record: dict[str, Any]) -> int:
    score = 0
    for field in ("platform", "country", "store", "store_id", "rating", "review", "translated_review", "customer", "review_time", "order_id", "order_total"):
        if _meaningful(record.get(field)):
            score += 4
    items = record.get("ordered_items")
    if isinstance(items, list) and items:
        score += 18 + min(len(items), 8)
        score += sum(
            1
            for item in items
            if isinstance(item, dict)
            and _meaningful(
                item.get("unit_price")
                or item.get("unitPrice")
                or item.get("price")
                or item.get("amount")
                or item.get("totalPrice")
                or item.get("productPrice")
                or item.get("goodsPrice")
                or item.get("salePrice")
                or item.get("finalPrice")
                or item.get("singlePrice")
                or item.get("subtotal")
                or item.get("order_total")
            )
        )
    if _meaningful(record.get("order_detail")):
        score += 18 + min(len(str(record.get("order_detail"))), 300) // 30
    images = record.get("image_urls")
    if isinstance(images, list) and images:
        score += 10 + min(len(images), 6)
    if isinstance(record.get("raw_json"), dict) and record["raw_json"]:
        score += min(len(record["raw_json"]), 20)
    if "exports\\runs" in str(record.get("source_file") or "") or "exports/runs" in str(record.get("source_file") or ""):
        score += 6
    return score


def _merge_review_records(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    primary, secondary = (
        (incoming, existing)
        if _review_completeness_score(incoming) > _review_completeness_score(existing)
        else (existing, incoming)
    )
    merged = dict(primary)
    for key, value in secondary.items():
        if not _meaningful(merged.get(key)) and _meaningful(value):
            merged[key] = value
            continue
        if key == "image_urls":
            urls = []
            for candidate in (_as_list(merged.get(key)), _as_list(value)):
                for url in candidate:
                    if url and url not in urls:
                        urls.append(url)
            merged[key] = urls
        elif key == "ordered_items":
            current = merged.get(key) if isinstance(merged.get(key), list) else []
            extra = value if isinstance(value, list) else []
            def item_score(items: list[Any]) -> int:
                total = 0
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    text = json.dumps(item, ensure_ascii=False)
                    has_name = _meaningful(
                        item.get("product")
                        or item.get("product_name")
                        or item.get("productName")
                        or item.get("goodsName")
                        or item.get("itemName")
                        or item.get("name")
                    )
                    has_qty = _meaningful(item.get("quantity") or item.get("qty") or item.get("count") or item.get("num"))
                    has_price = _meaningful(
                        item.get("unit_price")
                        or item.get("unitPrice")
                        or item.get("price")
                        or item.get("amount")
                        or item.get("totalPrice")
                        or item.get("productPrice")
                        or item.get("goodsPrice")
                        or item.get("salePrice")
                        or item.get("finalPrice")
                        or item.get("singlePrice")
                        or item.get("subtotal")
                        or item.get("order_total")
                    )
                    total += (2 if has_name else 0) + (3 if has_qty else 0) + (5 if has_price else 0)
                    if re.search(r"(商品|產品|产品|饮品|飲品|goods|product|item|price|单价|單價)", text, flags=re.IGNORECASE):
                        total += 2
                return total
            if item_score(extra) > item_score(current) or (item_score(extra) == item_score(current) and len(extra) > len(current)):
                merged[key] = extra
        elif key == "order_detail" and len(str(value or "")) > len(str(merged.get(key) or "")):
            merged[key] = value
        elif key == "raw_json" and isinstance(merged.get(key), dict) and isinstance(value, dict):
            merged[key] = {**value, **merged[key]}
    merged["has_order"] = bool(merged.get("order_id") or _meaningful(merged.get("order_detail")) or merged.get("ordered_items"))
    merged["has_image"] = bool(merged.get("image_urls"))
    return merged


def _is_placeholder_review(record: dict[str, Any]) -> bool:
    review_text = _clean_identity_text(record.get("review") or record.get("translated_review"), 80)
    if review_text not in {"no data", "暂无", "暂无数据", "empty", "n/a"}:
        return False
    return not (
        _meaningful(record.get("order_id"))
        or _meaningful(record.get("order_detail"))
        or bool(record.get("ordered_items"))
        or bool(record.get("image_urls"))
    )


def _is_non_review_page_text(record: dict[str, Any]) -> bool:
    text = str(record.get("review") or record.get("translated_review") or "").strip()
    if not text:
        return False
    platform = canonical_platform(str(record.get("platform") or "")).lower()
    has_core_review_evidence = any(
        [
            _meaningful(record.get("rating")),
            _meaningful(record.get("review_time")),
            _meaningful(record.get("order_id")),
        ]
    )
    if platform == "dianping" and not (
        _meaningful(record.get("rating")) or _meaningful(record.get("review_time"))
    ):
        return True
    site_chrome_patterns = (
        r"商户服务|关于我们|请登录|登录/注册|去\s*APP\s*查看更多内容|查看更多内容|美食",
        r"中国最贵的将军墓|上千瓶茅台|网红大滑梯|这个商场惊见",
        r"privacy policy|terms of use|sign in|log in|register|download app",
    )
    return (not has_core_review_evidence) and any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in site_chrome_patterns)


def _read_real_reviews(limit: int = 200) -> list[dict[str, Any]]:
    records_by_key: dict[str, dict[str, Any]] = {}
    fuzzy_buckets: dict[str, list[str]] = {}
    files = []
    if EXPORT_DIR.exists():
        files.extend(sorted(RUNS_DIR.glob("*/normalized_reviews.jsonl"), key=lambda p: p.stat().st_mtime, reverse=True) if RUNS_DIR.exists() else [])
        files.extend(sorted([p for p in EXPORT_DIR.rglob("*.json") if "runs" not in p.parts], key=lambda p: p.stat().st_mtime, reverse=True))
    for path in files:
        try:
            if path.suffix == ".jsonl":
                lines = path.read_text(encoding="utf-8-sig", errors="replace").splitlines()
                payload = {"run_id": path.parent.name}
                rows = [json.loads(line) for line in lines if line.strip()]
            else:
                payload, rows = _extract_reviews_from_payload(json.loads(path.read_text(encoding="utf-8-sig", errors="replace")))
            for index, raw in enumerate(rows, start=1):
                record = _normalize_ui_review(raw, payload, path, index)
                if not any(record.get(field) for field in ("review", "order_id", "order_detail", "ordered_items", "image_urls")):
                    continue
                if _is_placeholder_review(record):
                    continue
                if _is_non_review_page_text(record):
                    continue
                key = _resolve_review_dedupe_key(records_by_key, fuzzy_buckets, record)
                if key in records_by_key:
                    records_by_key[key] = _merge_review_records(records_by_key[key], record)
                else:
                    records_by_key[key] = record
                    _remember_review_dedupe_key(fuzzy_buckets, record, key)
        except Exception:
            continue
        if len(records_by_key) >= limit * 3:
            break
    records = list(records_by_key.values())
    return sorted(records, key=_review_sort_key, reverse=True)[:limit]


def _normalize_region_name(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    replacements = [
        ("香港特别行政区", "中国香港"),
        ("香港", "中国香港"),
        ("澳門特別行政區", "中国澳门"),
        ("澳门特别行政区", "中国澳门"),
        ("澳門", "中国澳门"),
        ("澳门", "中国澳门"),
        ("台灣地區", "中国台湾"),
        ("台湾地区", "中国台湾"),
        ("台灣", "中国台湾"),
        ("台湾", "中国台湾"),
        ("China Hong Kong", "China Hong Kong"),
        ("Hong Kong", "China Hong Kong"),
        ("Macao", "China Macau"),
        ("Macau", "China Macau"),
        ("Taiwan", "China Taiwan"),
    ]
    normalized = text
    for src, dst in replacements:
        normalized = normalized.replace(src, dst)
    normalized = re.sub(r"(中国)+香港", "中国香港", normalized)
    normalized = re.sub(r"(中国)+澳门", "中国澳门", normalized)
    normalized = re.sub(r"(中国)+台湾", "中国台湾", normalized)
    normalized = re.sub(r"(China\s+)+Hong Kong", "China Hong Kong", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"(China\s+)+Macau", "China Macau", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"(China\s+)+Taiwan", "China Taiwan", normalized, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", normalized).strip()


def _safe_float(value: Any) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def _resolve_directory_for_listing(path_value: str) -> Path:
    text = str(path_value or "").strip()
    if not text:
        candidate = ROOT
    else:
        raw = Path(text).expanduser()
        candidate = raw if raw.is_absolute() else (ROOT / raw)
    candidate = candidate.resolve(strict=False)
    probe = candidate
    while not probe.exists():
        if probe.parent == probe:
            break
        probe = probe.parent
    if probe.exists() and probe.is_file():
        probe = probe.parent
    if probe.exists() and probe.is_dir():
        return probe
    return ROOT


def _display_path(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT)).replace("\\", "/")
    except Exception:
        return str(path)


def _knowledge_index_path() -> Path:
    return KNOWLEDGE_DIR / "index.json"


def _load_knowledge_index() -> list[dict[str, Any]]:
    KNOWLEDGE_DIR.mkdir(parents=True, exist_ok=True)
    index_path = _knowledge_index_path()
    if not index_path.exists():
        return []
    try:
        payload = json.loads(index_path.read_text(encoding="utf-8"))
        if isinstance(payload, list):
            return [item for item in payload if isinstance(item, dict)]
    except Exception:
        return []
    return []


def _save_knowledge_index(entries: list[dict[str, Any]]) -> None:
    KNOWLEDGE_DIR.mkdir(parents=True, exist_ok=True)
    index_path = _knowledge_index_path()
    index_path.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")


def _decode_text_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "big5", "latin-1"):
        try:
            return data.decode(encoding)
        except Exception:
            continue
    return data.decode("utf-8", errors="replace")


def _extract_knowledge_file_text(filename: str, data: bytes) -> str:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".txt", ".md", ".json", ".csv", ".log"}:
        return _decode_text_bytes(data)
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from io import BytesIO
            from openpyxl import load_workbook  # type: ignore

            workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in workbook.worksheets[:8]:
                rows.append(f"# Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    values = [str(value).strip() for value in row if value not in (None, "")]
                    if values:
                        rows.append(" | ".join(values))
                    if len(rows) >= 4000:
                        break
            return "\n".join(rows)
        except Exception as exc:
            return f"[xlsx_parse_failed: {exc}]\n" + _decode_text_bytes(data[:200000])
    if suffix == ".docx":
        try:
            from io import BytesIO
            import docx  # type: ignore

            document = docx.Document(BytesIO(data))
            return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip())
        except Exception as exc:
            return f"[docx_parse_failed: {exc}]\n" + _decode_text_bytes(data[:200000])
    if suffix == ".pdf":
        try:
            from io import BytesIO
            try:
                from pypdf import PdfReader  # type: ignore
            except Exception:
                from PyPDF2 import PdfReader  # type: ignore

            reader = PdfReader(BytesIO(data))
            pages = []
            for page in reader.pages[:60]:
                pages.append(page.extract_text() or "")
            return "\n".join(text for text in pages if text.strip())
        except Exception as exc:
            return f"[pdf_parse_failed: {exc}]\n" + _decode_text_bytes(data[:200000])
    return _decode_text_bytes(data)


def _persist_knowledge_entry(name: str, content: str, source_type: str, tags: list[Any] | None = None) -> dict[str, Any]:
    clean_name = str(name or "").strip() or "knowledge_note"
    clean_content = str(content or "").strip()
    if not clean_content:
        raise ValueError("content is required")
    safe_name = re.sub(r"[^0-9A-Za-z_\-\u4e00-\u9fff]+", "_", clean_name).strip("_") or "knowledge_note"
    entry_id = uuid4().hex[:12]
    file_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{entry_id}_{safe_name}.md"
    file_path = KNOWLEDGE_DIR / file_name
    KNOWLEDGE_DIR.mkdir(parents=True, exist_ok=True)
    file_path.write_text(clean_content, encoding="utf-8")

    snippet = clean_content.replace("\r", " ").replace("\n", " ")[:320]
    now_iso = datetime.now().isoformat(timespec="seconds")
    entry = {
        "id": entry_id,
        "name": clean_name,
        "file": str(file_path.relative_to(ROOT)).replace("\\", "/"),
        "snippet": snippet,
        "tags": [str(item) for item in (tags or []) if str(item).strip()],
        "source_type": source_type,
        "created_at": now_iso,
        "updated_at": now_iso,
    }
    entries = [item for item in _load_knowledge_index() if str(item.get("id") or "") != entry_id]
    entries.append(entry)
    _save_knowledge_index(entries)
    EVENT_BUS.publish("success", "Knowledge base updated", f"Added knowledge entry: {clean_name}", {"entry_id": entry_id})
    return entry


def _knowledge_prompt_context(limit: int = 6, max_chars: int = 6000) -> str:
    entries = sorted(_load_knowledge_index(), key=lambda item: str(item.get("updated_at") or ""), reverse=True)[: max(1, limit)]
    blocks: list[str] = []
    consumed = 0
    for entry in entries:
        file_rel = str(entry.get("file") or "")
        file_path = ROOT / file_rel if file_rel else Path()
        snippet = str(entry.get("snippet") or "")
        if file_rel and file_path.exists():
            try:
                snippet = file_path.read_text(encoding="utf-8", errors="replace")[:1200]
            except Exception:
                snippet = str(entry.get("snippet") or "")
        text = f"[{entry.get('name') or file_rel}] {snippet}".strip()
        if not text:
            continue
        if consumed + len(text) > max_chars:
            remaining = max_chars - consumed
            if remaining <= 0:
                break
            text = text[:remaining]
        consumed += len(text)
        blocks.append(text)
        if consumed >= max_chars:
            break
    return "\n\n".join(blocks)


def _platform_fallback_steps(platform: str) -> list[str]:
    p = canonical_platform(platform)
    common = [
        "检查账号登录态、Cookie 和 MFA 是否过期。",
        "切换到只读路径（列表页/详情弹窗），避免点击回复或保存按钮。",
        "将时间范围先缩小到近7天，确认接口可读后再扩展。",
        "开启一次 Dry Run，保留 run_id 与 checkpoint 用于重试。",
    ]
    platform_specific: dict[str, list[str]] = {
        "google_maps": [
            "优先使用公开评论 DOM + 展开译文；若地图加载异常，切换代理节点后重试。",
            "分页滚动失败时降低滚动节奏并拉长等待时间（反爬保护）。",
        ],
        "grabfood": [
            "先确认进入 Portal 后再进 Feedback/Ratings 页面，避免跳转到营销页。",
            "门店筛选和时间筛选分步执行，筛选后等待表格稳定再抓取。",
        ],
        "hungry_panda": [
            "先进入 dataCenter 确认已登录，再进入 Orders → Ratings & Reviews。",
            "若分店切换失败，回到 Branch Management 后重新点 Branch Page。",
        ],
        "mfood": [
            "先进入订单管理 → 外卖评价，检查时间筛选器与分页是否可见。",
            "订单详情与查看图片只读打开后立即关闭，避免触发写操作。",
        ],
        "dianping": [
            "优先使用门店公开评论页 DOM 抽取，遇反爬时降低翻页频率并重试。",
            "仅采集评论正文/评分/时间/图片，订单详情字段标记为不适用。",
        ],
        "uber_eats": [
            "先确认商家账号会话可用，再进入对应门店反馈页；登录失败时触发人工门。",
            "优先读取评论列表与订单标识，订单详情只读展开后采集单价与商品项。",
        ],
        "aomi": [
            "确认 portal_url 可访问且已进入评价页；若缺失需先补充平台入口链接。",
            "优先按近7天抓取并验证详情弹窗结构，再扩展批量任务。",
        ],
    }
    return platform_specific.get(p, []) + common


def _diagnose_platform_connectivity(platform: str, runs: list[dict[str, Any]]) -> dict[str, Any]:
    filtered = [run for run in runs if canonical_platform(str(run.get("platform") or "")) == canonical_platform(platform)]
    if not filtered:
        return {
            "status": "unknown",
            "summary": "no run history for platform",
            "error_count_24h": 0,
            "review_count_24h": 0,
            "latest_run": {},
            "suggested_actions": _platform_fallback_steps(platform),
        }
    latest = filtered[0]
    error_count = sum(int(item.get("error_count") or 0) for item in filtered[:24])
    review_count = sum(int(item.get("review_count") or 0) for item in filtered[:24])
    latest_error = int(latest.get("error_count") or 0)
    latest_reviews = int(latest.get("review_count") or 0)
    if latest_error <= 0 and latest_reviews > 0:
        status = "ok"
        summary = "latest run successful"
    elif latest_error > 0 and latest_reviews <= 0:
        status = "failed"
        summary = "latest run failed without collected reviews"
    else:
        status = "degraded"
        summary = "partial collection; some errors detected"
    return {
        "status": status,
        "summary": summary,
        "error_count_24h": error_count,
        "review_count_24h": review_count,
        "latest_run": latest,
        "suggested_actions": _platform_fallback_steps(platform),
    }


async def _build_ai_remediation(platform: str, region: str, diagnosis: dict[str, Any]) -> dict[str, Any]:
    prompt = (
        "你是外卖平台采集系统的生产支持工程师。"
        "请根据以下诊断信息给出“实时处置方案”，仅输出 JSON，不要Markdown。\n"
        "字段要求：summary, urgency, root_causes[], immediate_actions[], retry_plan{attempts,backoff_minutes}, "
        "manual_gate_when[], verification_checks[]。\n"
        f"平台: {platform}\n"
        f"区域: {region or '-'}\n"
        f"诊断: {json.dumps(diagnosis, ensure_ascii=False)}\n"
    )
    text = await chat_with_configured_provider(prompt, max_tokens=800)
    candidate = text.strip()
    start = candidate.find("{")
    end = candidate.rfind("}")
    if start >= 0 and end > start:
        candidate = candidate[start : end + 1]
    parsed = json.loads(candidate)
    if isinstance(parsed, dict):
        return parsed
    raise ValueError("ai_response_not_object")


_KEYWORD_NOISE_WORDS = {
    "the", "and", "for", "with", "that", "this", "was", "were", "have", "has", "from", "very", "just", "but", "you",
    "your", "our", "their", "they", "been", "delivery", "drink", "review", "comment", "order", "service", "food",
    "local", "guide", "reviews", "photos", "photo", "ago", "new", "star", "stars", "rating", "rated", "minutes",
    "minute", "hours", "hour", "days", "day", "item", "items", "qty", "quantity", "price", "subtotal", "total",
    "评论", "評論", "則評論", "则评论", "配送", "门店", "客服", "这个", "那个", "我们", "你们", "他们", "因为", "但是", "没有",
    "分钟", "分鐘", "分钟前", "分鐘前", "小时", "小時", "小时前", "小時前", "天前", "新", "照片", "评分", "星", "颗星", "顆星",
    "订单", "订单号", "訂單", "商品", "规格", "規格", "数量", "數量", "单价", "單價", "价格", "價格", "小计", "小計",
    "标准", "標準", "甜度", "茶底", "调整", "調整", "状态", "狀態", "推荐", "推薦", "冰沙", "少少少甜", "去茶底",
    "条评价", "條評價", "則評價", "则评价", "在地嚮導", "本地向导", "張相片", "张照片", "週前", "周前",
    "您好", "我们深表歉意", "我们深感抱歉", "感謝您", "谢谢您", "app", "heytea", "tea", "there", "added",
    "making", "drinking", "even", "though", "requested", "bubble", "straw", "impossible", "pretty", "good",
    "hey", "sam", "chen", "nguyen", "此致", "好的", "然而", "而且", "个字", "好吧",
}

_BUSINESS_KEYWORD_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("缺少吸管/餐具", ("straw", "吸管", "餐具", "accessories", "utensil")),
    ("漏单少件", ("少给", "少了", "漏", "missing", "missed", "只给", "只給", "少杯", "少一杯", "少两杯", "少兩杯")),
    ("杯型/规格不符", ("大杯", "小杯", "杯子大小", "大小不一致", "wrong size", "size", "规格不符", "規格不符")),
    ("等待过久", ("等了", "等待", "太慢", "慢", "delay", "late", "wait", "超时", "超時")),
    ("包装问题", ("包装", "包裝", "漏洒", "漏灑", "spill", "spilled", "seal", "袋子", "破损", "破損")),
    ("口味正向", ("好喝", "味道很好", "味道很棒", "太棒了", "好茶", "tasty", "delicious", "love", "nice")),
    ("口味/甜度问题", ("太甜", "不甜", "没味", "沒有味", "no taste", "sweet", "bitter", "underwhelming", "不好喝")),
    ("服务态度问题", ("服务", "服務", "态度", "態度", "rude", "staff", "employee", "客服")),
    ("价格/性价比", ("贵", "貴", "pricey", "expensive", "worth", "性价比", "性價比")),
    ("食品安全风险", ("异物", "異物", "发霉", "發霉", "变质", "變質", "hair", "mold", "spoiled", "bug", "吐口水")),
    ("个性化需求/备注", ("生日", "蜡烛", "蠟燭", "candle", "birthday", "备注", "備註", "request")),
    ("商品热度/复购", ("每次都点", "每次都點", "必点", "必點", "go to", "go-to", "favorite", "favourite", "常点", "常點")),
)


def _clean_keyword_review_text(review: dict[str, Any]) -> str:
    for raw in (review.get("review"), review.get("translated_review")):
        text = str(raw or "").strip()
        if not text or text in {"-", "None", "none", "null", "No data", "暂无", "暂无数据"}:
            continue
        text = re.sub(r"[\ue000-\uf8ff]", " ", text)
        text = re.sub(r"\b\d+\s*(?:reviews?|photos?|stars?|minutes?|hours?|days?)\b", " ", text, flags=re.IGNORECASE)
        text = re.sub(r"\d+\s*(?:则评论|則評論|张照片|張照片|分钟前|分鐘前|小时前|小時前|天前)", " ", text)
        text = re.sub(r"\b(?:local guide|new|ago)\b", " ", text, flags=re.IGNORECASE)
        text = re.sub(r"(?:订单号|訂單號|order\s*(?:id|no\.?|number|#))\s*[:：#]?\s*[A-Z]{0,6}\d{6,}", " ", text, flags=re.IGNORECASE)
        text = re.sub(r"\s+", " ", text).strip()
        if len(text) >= 2 and not re.fullmatch(r"[\d\s:：/\-.,，。]+", text):
            return text
    return ""


def _extract_keywords(reviews: list[dict[str, Any]], top_n: int = 15) -> list[dict[str, Any]]:
    business_counts: dict[str, int] = {}
    counts: dict[str, int] = {}
    for review in reviews:
        text = _clean_keyword_review_text(review).lower()
        if not text:
            continue
        for label, needles in _BUSINESS_KEYWORD_RULES:
            if any(needle.lower() in text for needle in needles):
                business_counts[label] = business_counts.get(label, 0) + 1
        en_tokens = re.findall(r"[a-z][a-z'-]{2,}", text)
        zh_tokens = re.findall(r"[\u4e00-\u9fff]{2,}", text)
        for token in en_tokens + zh_tokens:
            token = token.strip()
            if (
                not token
                or token in _KEYWORD_NOISE_WORDS
                or token.isdigit()
                or re.fullmatch(r"\d+[a-z]*", token, flags=re.IGNORECASE)
                or re.search(r"(分钟|分鐘|小时|小時|天前|週前|周前|评论|評論|評價|评价|照片|相片|向导|嚮導|订单|訂單|规格|規格|数量|數量|价格|價格|歉意|抱歉)", token)
            ):
                continue
            counts[token] = counts.get(token, 0) + 1
    business_top = sorted(business_counts.items(), key=lambda item: item[1], reverse=True)
    return [{"keyword": word, "count": count} for word, count in business_top[: max(1, top_n)]]


def _cluster_reviews(reviews: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets = {
        "product_quality": ["异物", "变质", "发霉", "mold", "spoiled", "hair", "bug"],
        "delivery_speed": ["慢", "超时", "delay", "late", "wait"],
        "service_attitude": ["服务", "态度", "rude", "客服", "complaint"],
        "packaging": ["包装", "漏", "spill", "seal", "cup"],
        "taste": ["口味", "味道", "tasty", "sweet", "bitter"],
    }
    out: list[dict[str, Any]] = []
    for cluster, words in buckets.items():
        matched = []
        for review in reviews:
            text = f"{review.get('review', '')} {review.get('translated_review', '')}".lower()
            if any(word.lower() in text for word in words):
                matched.append(review)
        out.append({"cluster": cluster, "count": len(matched)})
    return out


def _parse_review_date_text(value: str) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    for token in (text[:10], text.replace("/", "-")[:10]):
        try:
            return datetime.fromisoformat(token).date()
        except Exception:
            continue
    return None


def _resolve_time_range_bounds(time_range: Any, fallback_days: int = 7) -> tuple[date, date, int]:
    safe_days = int(getattr(time_range, "days", 0) or fallback_days or 7)
    safe_days = max(1, min(safe_days, 30))
    today = datetime.now().date()
    start = (
        date.fromisoformat(str(getattr(time_range, "start_date", "")).strip())
        if getattr(time_range, "start_date", "")
        else (today - timedelta(days=safe_days - 1))
    )
    end = (
        date.fromisoformat(str(getattr(time_range, "end_date", "")).strip())
        if getattr(time_range, "end_date", "")
        else today
    )
    if end < start:
        start, end = end, start
    span = max(1, (end - start).days + 1)
    return start, end, span


def _daily_volume_series(
    reviews: list[dict[str, Any]],
    days: int,
    start_date: date | None = None,
    end_date: date | None = None,
) -> list[dict[str, Any]]:
    counts: dict[str, int] = {}
    for review in reviews:
        key = str(review.get("review_time") or "")[:10].replace("/", "-")
        if re.match(r"^\d{4}-\d{2}-\d{2}$", key):
            counts[key] = counts.get(key, 0) + 1
    if start_date and end_date:
        start = min(start_date, end_date)
        end = max(start_date, end_date)
        span = max(1, (end - start).days + 1)
    else:
        today = datetime.now().date()
        span = max(1, days)
        end = today
        start = today - timedelta(days=span - 1)
    result = []
    for offset in range(span):
        day = (start + timedelta(days=offset)).isoformat()
        result.append({"date": day, "count": counts.get(day, 0)})
    return result


def _platform_volume_series(reviews: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counts: dict[str, int] = {}
    for review in reviews:
        platform = canonical_platform(str(review.get("platform") or ""))
        if not platform:
            continue
        counts[platform] = counts.get(platform, 0) + 1
    return [{"platform": key, "count": value} for key, value in sorted(counts.items(), key=lambda item: item[1], reverse=True)]


def _lifecycle_series(reviews: list[dict[str, Any]], days: int) -> list[dict[str, Any]]:
    now = datetime.now().date()
    buckets = [
        ("fresh_0_1d", 0, 1),
        ("watch_2_3d", 2, 3),
        ("manage_4_7d", 4, 7),
        ("followup_8_14d", 8, 14),
        ("archive_15_30d", 15, max(15, days)),
    ]
    counts = {key: 0 for key, _start, _end in buckets}
    for review in reviews:
        review_time = str(review.get("review_time") or "")
        parsed_date = None
        for token in (review_time[:10], review_time.replace("/", "-")[:10]):
            try:
                parsed_date = datetime.fromisoformat(token).date()
                break
            except Exception:
                continue
        if not parsed_date:
            continue
        age = max(0, (now - parsed_date).days)
        for key, start, end in buckets:
            if start <= age <= end:
                counts[key] += 1
                break
    return [{"stage": key, "count": counts[key]} for key, _start, _end in buckets]


def _top_risk_reviews(reviews: list[dict[str, Any]], limit: int = 20) -> list[dict[str, Any]]:
    scored: list[tuple[int, dict[str, Any]]] = []
    for review in reviews:
        text = f"{review.get('review', '')} {review.get('translated_review', '')}"
        score = 0
        rating = _safe_float(review.get("rating"))
        if rating and rating <= 2:
            score += 3
        if re.search(r"异物|发霉|变质|腹泻|投诉|slow|late|hair|mold|spoiled|rude", text, re.IGNORECASE):
            score += 2
        if review.get("has_image"):
            score += 1
        if score > 0:
            scored.append((score, review))
    scored.sort(key=lambda pair: pair[0], reverse=True)
    return [row for _score, row in scored[: max(1, limit)]]


async def _build_ai_insight(metrics: dict[str, Any], clusters: list[dict[str, Any]], keywords: list[dict[str, Any]]) -> dict[str, Any]:
    prompt = (
        "你是外卖评论数据分析专家。根据下面实时统计，输出JSON："
        "字段包括 summary, key_findings[], root_causes[], actions[], risk_level。\n"
        f"metrics={json.dumps(metrics, ensure_ascii=False)}\n"
        f"clusters={json.dumps(clusters, ensure_ascii=False)}\n"
        f"keywords={json.dumps(keywords, ensure_ascii=False)}\n"
        "要求：每条建议可执行、可验证。"
    )
    text = await chat_with_configured_provider(prompt, max_tokens=1000)
    candidate = text.strip()
    start = candidate.find("{")
    end = candidate.rfind("}")
    if start >= 0 and end > start:
        candidate = candidate[start : end + 1]
    parsed = json.loads(candidate)
    if isinstance(parsed, dict):
        return parsed
    raise ValueError("ai_insight_not_object")


async def _build_ai_insight_v2(
    metrics: dict[str, Any],
    clusters: list[dict[str, Any]],
    keywords: list[dict[str, Any]],
    reviews: list[dict[str, Any]],
    days: int,
) -> dict[str, Any]:
    knowledge_context = _knowledge_prompt_context(limit=8, max_chars=7000)
    review_samples = [
        {
            "platform": row.get("platform"),
            "country": row.get("country"),
            "store": row.get("store"),
            "rating": row.get("rating"),
            "review_time": row.get("review_time"),
            "review": str(row.get("review") or "")[:240],
            "translated_review": str(row.get("translated_review") or "")[:240],
            "order_id": row.get("order_id"),
            "has_image": bool(row.get("has_image")),
            "has_order": bool(row.get("has_order")),
        }
        for row in reviews[:60]
    ]
    prompt = (
        "你是连锁餐饮评论治理专家。请基于真实评论数据输出 JSON，不要 Markdown。\n"
        "字段必须包含：summary, key_findings[], root_causes[], actions[], risk_level, "
        "trend_observation, lifecycle_stage, complaint_clusters[], food_safety_issues[].\n"
        f"time_window_days={days}\n"
        f"metrics={json.dumps(metrics, ensure_ascii=False)}\n"
        f"clusters={json.dumps(clusters, ensure_ascii=False)}\n"
        f"keywords={json.dumps(keywords, ensure_ascii=False)}\n"
        f"review_samples={json.dumps(review_samples, ensure_ascii=False)}\n"
        f"knowledge_base={knowledge_context or 'N/A'}\n"
        "要求：所有结论必须可追溯，不允许臆造。"
    )
    text = await chat_with_configured_provider(prompt, max_tokens=1600)
    candidate = text.strip()
    start = candidate.find("{")
    end = candidate.rfind("}")
    if start >= 0 and end > start:
        candidate = candidate[start : end + 1]
    parsed = json.loads(candidate)
    if isinstance(parsed, dict):
        return parsed
    raise ValueError("ai_insight_not_object")


@nicegui_app.get("/api/unified/status")
def api_unified_status() -> dict:
    registry = _load_safe_registry()
    exports = [
        {
            "name": path.name,
            "path": str(path.relative_to(ROOT)),
            "bytes": path.stat().st_size,
            "mtime": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
        }
        for path in sorted(
            [p for p in EXPORT_DIR.rglob("*") if p.is_file()] if EXPORT_DIR.exists() else [],
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        )[:20]
    ]
    return {
        "ok": True,
        "now": datetime.now().isoformat(timespec="seconds"),
        "store_count": registry["store_count"],
        "platform_counts": registry["platform_counts"],
        "platforms": {key: capability.to_dict() for key, capability in PLATFORM_CAPABILITIES.items()},
        "tasks": [path.name for path in _task_files()],
        "exports": exports,
        "runs_api": "/api/unified/runs",
        "diagnose_api": "/api/unified/platform-diagnose",
        "coordinator": COORDINATOR.snapshot(),
        "monitor": SYNC_MONITOR.status(),
        "safety": {
            "mode": "read_only",
            "allowed": ["login", "navigate", "search", "filter", "open_detail", "export", "dry_run"],
            "denied": ["reply", "save", "submit", "delete", "confirm", "payment", "modify_store_config"],
        },
    }


@nicegui_app.get("/api/unified/stores")
def api_unified_stores(country: str = "", platform: str = "", limit: int = 300) -> dict:
    registry = _load_safe_registry()
    canonical = canonical_platform(platform) if platform else ""
    country_filter = _normalize_region_name(str(country or ""))
    stores = []
    for store in registry["stores"]:
        if country_filter and country_filter not in _normalize_region_name(str(store.get("country", ""))):
            continue
        if canonical and canonical not in {canonical_platform(key) for key in store.get("platforms", {})}:
            continue
        stores.append(store)
    return {"ok": True, "count": len(stores), "stores": stores[: max(1, min(limit, 1000))]}


@nicegui_app.get("/api/unified/tasks")
def api_unified_tasks() -> dict:
    tasks = []
    for path in _task_files():
        try:
            task = load_task(path)
            tasks.append({"name": path.name, "task": task.to_dict()})
        except Exception as exc:
            tasks.append({"name": path.name, "error": str(exc)})
    return {"ok": True, "tasks": tasks}


@nicegui_app.get("/api/unified/runs")
def api_unified_runs(limit: int = 50) -> dict:
    runs = []
    if RUNS_DIR.exists():
        for path in sorted([item for item in RUNS_DIR.iterdir() if item.is_dir()], key=lambda item: item.stat().st_mtime, reverse=True)[
            : max(1, min(limit, 200))
        ]:
            checkpoint = path / "checkpoint.json"
            quality_report = path / "quality_report.json"
            normalized = path / "normalized_reviews.jsonl"
            item = {
                "run_id": path.name,
                "run_dir": str(path.relative_to(ROOT)),
                "has_checkpoint": checkpoint.exists(),
                "has_normalized_reviews": normalized.exists(),
                "has_quality_report": quality_report.exists(),
                "updated_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
            }
            if checkpoint.exists():
                try:
                    data = json.loads(checkpoint.read_text(encoding="utf-8"))
                    item["last_stage"] = data.get("last_stage", "")
                    checkpoints = data.get("checkpoints") or []
                    started = next((cp for cp in checkpoints if cp.get("stage") == "started"), None)
                    task_info = ((started or {}).get("payload") or {}).get("task") or {}
                    if isinstance(task_info, dict):
                        item["platform"] = task_info.get("platform", "")
                        item["account"] = task_info.get("account", "") or task_info.get("country", "")
                    errors = []
                    for cp in checkpoints:
                        payload = cp.get("payload") if isinstance(cp, dict) else {}
                        if not isinstance(payload, dict):
                            continue
                        for key in ("error", "stderr", "message"):
                            value = payload.get(key)
                            if value:
                                errors.append(str(value))
                        payload_errors = payload.get("errors")
                        if isinstance(payload_errors, list):
                            errors.extend(str(value) for value in payload_errors if str(value).strip())
                        elif payload_errors:
                            errors.append(str(payload_errors))
                        result = payload.get("result")
                        if isinstance(result, dict):
                            result_errors = result.get("errors")
                            if isinstance(result_errors, list):
                                errors.extend(str(value) for value in result_errors if str(value).strip())
                            elif result_errors:
                                errors.append(str(result_errors))
                            if result.get("error"):
                                errors.append(str(result.get("error")))
                    if errors:
                        deduped_errors = []
                        seen_errors = set()
                        for error in errors:
                            short = re.sub(r"\s+", " ", str(error)).strip()
                            if not short or short in seen_errors:
                                continue
                            seen_errors.add(short)
                            deduped_errors.append(short[:1200])
                        item["errors"] = deduped_errors[:8]
                except Exception:
                    item["last_stage"] = "checkpoint_unreadable"
            if quality_report.exists():
                try:
                    report = json.loads(quality_report.read_text(encoding="utf-8"))
                    item["review_count"] = report.get("review_count", 0)
                    item["field_completeness"] = report.get("field_completeness", 0)
                    item["detail_coverage"] = report.get("detail_coverage", 0)
                    item["error_count"] = report.get("error_count", 0)
                except Exception:
                    item["quality_error"] = "quality_report_unreadable"
            platform_key = canonical_platform(str(item.get("platform") or ""))
            missing_artifact = str(item.get("last_stage") or "").lower().startswith("finish") and not item.get("has_normalized_reviews")
            error_count = int(item.get("error_count") or 0)
            review_count = int(item.get("review_count") or 0)
            if missing_artifact or error_count > 0:
                item["health_status"] = "failed" if missing_artifact and review_count <= 0 else "partial"
                item["retry_suggestions"] = _platform_fallback_steps(platform_key)[:5] if platform_key else []
            elif str(item.get("last_stage") or "").lower() == "started":
                item["health_status"] = "running"
            elif review_count == 0 and item.get("has_quality_report"):
                item["health_status"] = "empty"
                item["retry_suggestions"] = [
                    "Verify the selected date range has merchant reviews.",
                    "Retry with visible browser mode if the platform may require session refresh.",
                    "Check whether the platform account is scoped to the expected stores.",
                ]
            else:
                item["health_status"] = "ok"
            runs.append(item)
    return {"ok": True, "count": len(runs), "runs": runs}


@nicegui_app.get("/api/unified/failures")
def api_unified_failures(limit: int = 50) -> dict:
    runs = list(api_unified_runs(limit=max(50, min(int(limit or 50) * 2, 200))).get("runs") or [])
    failures = []
    for run in runs:
        status = str(run.get("health_status") or "").lower()
        errors = run.get("errors") if isinstance(run.get("errors"), list) else []
        if status not in {"failed", "partial", "empty"} and not errors:
            continue
        platform = canonical_platform(str(run.get("platform") or ""))
        failures.append(
            {
                "run_id": run.get("run_id", ""),
                "platform": platform,
                "platform_label": run.get("platform", ""),
                "account": run.get("account", ""),
                "updated_at": run.get("updated_at", ""),
                "status": status or "failed",
                "review_count": run.get("review_count", 0),
                "error_count": run.get("error_count", len(errors)),
                "errors": errors[:5],
                "retry_suggestions": (run.get("retry_suggestions") or _platform_fallback_steps(platform))[:6],
                "run_dir": run.get("run_dir", ""),
            }
        )
        if len(failures) >= max(1, min(int(limit or 50), 100)):
            break
    events = [
        event
        for event in EVENT_BUS.list_since(max(0, EVENT_BUS.latest_id() - 120))
        if str(event.get("level") or "").lower() == "error"
    ][-20:]
    return {"ok": True, "count": len(failures), "failures": failures, "error_events": events}


@nicegui_app.post("/api/unified/platform-diagnose")
async def api_unified_platform_diagnose(request: Request) -> dict:
    body = await request.json()
    platform = canonical_platform(str((body or {}).get("platform") or ""))
    region = _normalize_region_name(str((body or {}).get("region") or (body or {}).get("country") or ""))
    if not platform:
        return {"ok": False, "error": "platform is required"}

    runs_payload = api_unified_runs(limit=160)
    runs = list(runs_payload.get("runs") or [])
    diagnosis = _diagnose_platform_connectivity(platform, runs)

    events = EVENT_BUS.list_since(max(0, EVENT_BUS.latest_id() - 120))
    recent_errors = []
    for event in reversed(events):
        level = str(event.get("level") or "").lower()
        message = f"{event.get('title', '')} {event.get('message', '')}".lower()
        payload_text = json.dumps(event.get("payload") or {}, ensure_ascii=False).lower()
        if level == "error" and (platform in message or platform in payload_text):
            recent_errors.append(event)
        if len(recent_errors) >= 10:
            break

    remediation = {
        "summary": diagnosis.get("summary", ""),
        "urgency": "high" if diagnosis.get("status") == "failed" else ("medium" if diagnosis.get("status") == "degraded" else "low"),
        "root_causes": [
            "登录态失效或地区门店切换失败",
            "页面结构变化（导航/筛选器/详情弹窗）",
            "接口限流或网络波动导致请求失败",
        ],
        "immediate_actions": diagnosis.get("suggested_actions", []),
        "retry_plan": {"attempts": 3, "backoff_minutes": [5, 15, 30]},
        "manual_gate_when": [
            "出现验证码/二次验证",
            "页面出现保存、回复、提交等写动作入口",
            "连续两次重试仍无法打开评论列表",
        ],
        "verification_checks": [
            "确认近7天评论数量 > 0",
            "确认订单详情字段可读（order_id/order_detail）",
            "确认图片 URL 字段非头像且可访问",
        ],
    }

    ai_enabled = bool((body or {}).get("use_ai", True))
    ai_error = ""

    ai_used = False
    if ai_enabled:
        try:
            ai_result = await _build_ai_remediation(platform, region, diagnosis)
            if isinstance(ai_result, dict) and ai_result:
                remediation = {**remediation, **ai_result}
                ai_used = True
        except Exception as exc:
            ai_error = str(exc)

    return {
        "ok": True,
        "platform": platform,
        "region": region,
        "diagnosis": diagnosis,
        "recent_errors": recent_errors,
        "remediation": remediation,
        "ai_used": ai_used,
        "ai_error": ai_error,
    }


@nicegui_app.get("/api/unified/reviews")
def api_unified_reviews(
    platform: str = "",
    country: str = "",
    store: str = "",
    days: int | None = 30,
    start_date: str = "",
    end_date: str = "",
    has_image: bool = False,
    has_order: bool = False,
    limit: int = 200,
) -> dict:
    try:
        if str(start_date or "").strip() and str(end_date or "").strip():
            time_range = validate_week_range(start_date=str(start_date).strip(), end_date=str(end_date).strip())
        else:
            safe_days = int(days) if days is not None else 30
            time_range = validate_week_range(days=safe_days)
    except ValueError as exc:
        return {"ok": False, "error": str(exc), "reviews": []}
    start_bound, end_bound, _span = _resolve_time_range_bounds(time_range, fallback_days=int(days or 30))
    platform_query = platform.lower().strip()
    country_query = _normalize_region_name(country).lower().strip()
    store_query = store.lower().strip()
    filtered = []
    for record in _read_real_reviews(limit=max(limit, 300)):
        if platform_query and platform_query not in str(record.get("platform", "")).lower():
            continue
        if country_query and country_query not in _normalize_region_name(str(record.get("country", ""))).lower():
            continue
        if store_query and store_query not in str(record.get("store", "")).lower():
            continue
        if has_image and not record.get("has_image"):
            continue
        if has_order and not record.get("has_order"):
            continue
        review_time = str(record.get("review_time") or "")
        parsed_date = _parse_review_date_text(review_time)
        if parsed_date and (parsed_date < start_bound or parsed_date > end_bound):
            continue
        filtered.append(record)
        if len(filtered) >= max(1, min(limit, 1000)):
            break
    return {
        "ok": True,
        "count": len(filtered),
        "time_range": time_range.__dict__,
        "source": "exports_only_real_records",
        "reviews": filtered,
    }


@nicegui_app.get("/api/unified/insight")
async def api_unified_insight(
    days: int | None = 7,
    limit: int = 1200,
    platform: str = "",
    start_date: str = "",
    end_date: str = "",
) -> dict:
    safe_days = 30 if int(days or 7) == 30 else 7
    payload = api_unified_reviews(
        platform=platform,
        days=safe_days,
        start_date=start_date,
        end_date=end_date,
        limit=max(300, min(limit, 3000)),
    )
    reviews = list(payload.get("reviews") or [])
    raw_range = payload.get("time_range") or {}
    range_start, range_end, range_days = _resolve_time_range_bounds(
        type("Range", (), {
            "start_date": str(raw_range.get("start_date") or ""),
            "end_date": str(raw_range.get("end_date") or ""),
            "days": int(raw_range.get("days") or safe_days),
        })(),
        fallback_days=safe_days,
    )
    days_for_series = max(1, min(range_days, 30))
    if not reviews:
        return {
            "ok": True,
            "days": days_for_series,
            "time_range": {
                "type": str(raw_range.get("type") or "last_days"),
                "days": days_for_series,
                "start_date": range_start.isoformat(),
                "end_date": range_end.isoformat(),
            },
            "metrics": {"review_count": 0, "risk_count": 0, "risk_index": 0, "platform_count": 0},
            "series": {"daily_volume": [], "platform_volume": [], "keywords": [], "clusters": [], "lifecycle": [], "risk_samples": []},
            "ai": {
                "summary": "当前时间范围内暂无可分析评论数据。",
                "key_findings": [],
                "root_causes": [],
                "actions": ["先执行采集任务，再刷新分析视图。"],
                "risk_level": "low",
                "trend_observation": "",
                "lifecycle_stage": "empty",
                "complaint_clusters": [],
                "food_safety_issues": [],
            },
            "knowledge_entries": _load_knowledge_index()[:10],
        }

    risk_reviews = 0
    for review in reviews:
        rating = _safe_float(review.get("rating"))
        text = f"{review.get('review', '')} {review.get('translated_review', '')}"
        if (rating > 0 and rating <= 2) or re.search(r"异物|发霉|变质|腹泻|投诉|slow|late|hair|mold|spoiled|rude", text, re.IGNORECASE):
            risk_reviews += 1

    keywords = _extract_keywords(reviews, top_n=20)
    clusters = _cluster_reviews(reviews)
    metrics = {
        "review_count": len(reviews),
        "risk_count": risk_reviews,
        "risk_index": round((risk_reviews / max(1, len(reviews))), 4),
        "platform_count": len({canonical_platform(str(item.get("platform") or "")) for item in reviews if item.get("platform")}),
    }
    series = {
        "daily_volume": _daily_volume_series(reviews, days=days_for_series, start_date=range_start, end_date=range_end),
        "platform_volume": _platform_volume_series(reviews),
        "keywords": keywords,
        "clusters": clusters,
        "lifecycle": _lifecycle_series(reviews, days=days_for_series),
        "risk_samples": _top_risk_reviews(reviews, limit=20),
    }
    fallback_ai = {
        "summary": f"共分析 {len(reviews)} 条评论，识别风险评论 {risk_reviews} 条。",
        "key_findings": [
            f"高频词前3：{', '.join(item['keyword'] for item in keywords[:3]) if keywords else '暂无'}",
            f"聚类Top：{max(clusters, key=lambda item: item['count'])['cluster'] if clusters else '暂无'}",
        ],
        "root_causes": [
            "高峰时段配送履约波动",
            "部分门店出品与包装一致性不足",
            "异常评论处理时效不稳定",
        ],
        "actions": [
            "优先处理近24小时低评分+带图评论，形成门店整改清单。",
            "对高风险关键词门店启用每日复采与复核。",
            "按平台建立重试策略：5/15/30分钟退避并记录失败证据。",
        ],
        "risk_level": "medium" if risk_reviews > 0 else "low",
    }

    ai_used = False
    ai_error = ""
    ai = fallback_ai
    try:
        ai = await _build_ai_insight_v2(metrics, clusters, keywords, reviews, days_for_series)
        ai_used = True
    except Exception as exc:
        ai_error = str(exc)

    return {
        "ok": True,
        "days": days_for_series,
        "time_range": {
            "type": str(raw_range.get("type") or "last_days"),
            "days": days_for_series,
            "start_date": range_start.isoformat(),
            "end_date": range_end.isoformat(),
        },
        "metrics": metrics,
        "series": series,
        "ai": ai,
        "ai_used": ai_used,
        "ai_error": ai_error,
        "knowledge_entries": _load_knowledge_index()[:10],
    }


def _compute_quality_metrics(reviews: list[dict[str, Any]], history: list[dict[str, Any]], days: int) -> dict[str, Any]:
    safe_days = 30 if int(days or 7) == 30 else 7
    start_date = datetime.now().date() - timedelta(days=safe_days - 1)
    required_fields = ("platform", "store", "review_time", "rating", "review")
    filled_slots = 0
    review_count = len(reviews)
    with_images = 0
    out_of_bounds = 0
    duplicate_rows = 0
    order_rows = 0
    order_with_detail = 0
    seen: set[str] = set()
    fuzzy_records: dict[str, dict[str, Any]] = {}
    fuzzy_buckets: dict[str, list[str]] = {}

    for review in reviews:
        snapshot = {
            "platform": str(review.get("platform") or "").strip(),
            "store": str(review.get("store") or "").strip(),
            "review_time": str(review.get("review_time") or "").strip(),
            "rating": str(review.get("rating") or "").strip(),
            "review": str(review.get("review") or review.get("translated_review") or "").strip(),
        }
        for field in required_fields:
            value = str(snapshot.get(field) or "").strip()
            if value and value not in {"-", "None", "none", "null"}:
                filled_slots += 1

        image_urls = review.get("image_urls")
        image_text = str(image_urls or "").strip()
        has_image = bool(review.get("has_image")) or (isinstance(image_urls, list) and len(image_urls) > 0) or (
            image_text and image_text != "-"
        )
        if has_image:
            with_images += 1

        review_time = str(review.get("review_time") or "").strip()
        parsed_date = None
        for token in (review_time[:10], review_time.replace("/", "-")[:10]):
            try:
                parsed_date = datetime.fromisoformat(token).date()
                break
            except Exception:
                pass
        if parsed_date and parsed_date < start_date:
            out_of_bounds += 1

        has_order_hint = bool(review.get("has_order")) or bool(str(review.get("order_id") or review.get("order_sn") or "").strip())
        if has_order_hint:
            order_rows += 1
            detail_text = str(review.get("order_detail") or "").strip()
            items_text = str(review.get("ordered_items_text") or "").strip()
            items = review.get("ordered_items")
            if (detail_text and detail_text != "-") or (items_text and items_text != "-") or (isinstance(items, list) and len(items) > 0):
                order_with_detail += 1

        dedupe_key = _resolve_review_dedupe_key(fuzzy_records, fuzzy_buckets, review)
        if dedupe_key in seen:
            duplicate_rows += 1
            fuzzy_records[dedupe_key] = _merge_review_records(fuzzy_records[dedupe_key], review)
        else:
            seen.add(dedupe_key)
            fuzzy_records[dedupe_key] = review
            _remember_review_dedupe_key(fuzzy_buckets, review, dedupe_key)

    manual_gate_count = 0
    total_errors = 0
    for entry in history:
        errors = entry.get("errors")
        if isinstance(errors, (list, tuple)):
            error_list = [str(item) for item in errors if str(item).strip()]
        elif errors:
            error_list = [str(errors)]
        else:
            error_list = []
        total_errors += len(error_list)
        if any(re.search(r"captcha|manual|blocked|forbidden|write|login", message, re.IGNORECASE) for message in error_list):
            manual_gate_count += 1

    slots = max(1, review_count * len(required_fields))
    return {
        "field_completion_rate": round((filled_slots / slots) * 100, 1),
        "detail_coverage": round((order_with_detail / max(1, order_rows)) * 100, 1),
        "image_coverage": round((with_images / max(1, review_count)) * 100, 1),
        "duplicate_rate": round((duplicate_rows / max(1, review_count)) * 100, 2),
        "out_of_bounds_count": int(out_of_bounds),
        "manual_gate_count": int(manual_gate_count),
        "total_errors": int(total_errors),
        "review_count": int(review_count),
    }


def _build_local_quality_insight(
    reviews: list[dict[str, Any]],
    days: int,
    start_date: date | None = None,
    end_date: date | None = None,
) -> dict[str, Any]:
    safe_days = 30 if int(days or 7) == 30 else 7
    risk_reviews = 0
    for review in reviews:
        rating = _safe_float(review.get("rating"))
        text = f"{review.get('review', '')} {review.get('translated_review', '')}"
        if (rating > 0 and rating <= 2) or re.search(r"异物|发霉|变质|腹泻|投诉|slow|late|hair|mold|spoiled|rude", text, re.IGNORECASE):
            risk_reviews += 1
    keywords = _extract_keywords(reviews, top_n=20)
    clusters = _cluster_reviews(reviews)
    metrics = {
        "review_count": len(reviews),
        "risk_count": risk_reviews,
        "risk_index": round((risk_reviews / max(1, len(reviews))), 4),
        "platform_count": len({canonical_platform(str(item.get("platform") or "")) for item in reviews if item.get("platform")}),
    }
    return {
        "ok": True,
        "days": safe_days,
        "metrics": metrics,
        "series": {
            "daily_volume": _daily_volume_series(reviews, days=safe_days, start_date=start_date, end_date=end_date),
            "platform_volume": _platform_volume_series(reviews),
            "keywords": keywords,
            "clusters": clusters,
            "lifecycle": _lifecycle_series(reviews, days=safe_days),
            "risk_samples": _top_risk_reviews(reviews, limit=20),
        },
        "ai": {
            "summary": f"实时分析共覆盖 {len(reviews)} 条评论，识别风险评论 {risk_reviews} 条。",
            "key_findings": [
                f"高频词前3：{', '.join(item['keyword'] for item in keywords[:3]) if keywords else '暂无'}",
                f"聚类Top：{max(clusters, key=lambda item: item['count'])['cluster'] if clusters else '暂无'}",
            ],
            "actions": [
                "优先复核低评分+带图评论，建立门店整改清单。",
                "对高风险关键词门店提高复采频率并安排复核。",
            ],
            "risk_level": "medium" if risk_reviews > 0 else "low",
            "trend_observation": "",
            "lifecycle_stage": "",
            "complaint_clusters": [item.get("cluster") for item in clusters[:5]],
            "food_safety_issues": [
                item.get("keyword")
                for item in keywords
                if any(token in str(item.get("keyword") or "") for token in ("异物", "发霉", "变质", "mold", "spoiled", "hair"))
            ][:6],
        },
        "ai_used": False,
        "ai_error": "",
    }


@nicegui_app.get("/api/unified/quality-report")
async def api_unified_quality_report(
    days: int | None = 7,
    limit: int = 1600,
    platform: str = "",
    start_date: str = "",
    end_date: str = "",
) -> dict:
    safe_days = 30 if int(days or 7) == 30 else 7
    safe_limit = max(300, min(int(limit or 1600), 4000))
    status = api_unified_status()
    reviews_payload = api_unified_reviews(
        platform=platform,
        days=safe_days,
        start_date=start_date,
        end_date=end_date,
        limit=safe_limit,
    )
    knowledge_payload = api_unified_knowledge(limit=30)
    settings_payload = api_unified_settings()
    reviews = list(reviews_payload.get("reviews") or [])
    history = list((status.get("coordinator") or {}).get("history") or [])
    raw_range = reviews_payload.get("time_range") or {}
    range_start, range_end, range_days = _resolve_time_range_bounds(
        type("Range", (), {
            "start_date": str(raw_range.get("start_date") or ""),
            "end_date": str(raw_range.get("end_date") or ""),
            "days": int(raw_range.get("days") or safe_days),
        })(),
        fallback_days=safe_days,
    )
    span_days = max(1, min(range_days, 30))
    metrics_quality = _compute_quality_metrics(reviews, history, span_days)
    insight_payload = _build_local_quality_insight(reviews, span_days, start_date=range_start, end_date=range_end)
    return {
        "ok": True,
        "days": span_days,
        "time_range": {
            "type": str(raw_range.get("type") or "last_days"),
            "days": span_days,
            "start_date": range_start.isoformat(),
            "end_date": range_end.isoformat(),
        },
        "last_updated": datetime.now().isoformat(timespec="seconds"),
        "platform": canonical_platform(platform) if platform else "",
        "metrics_quality": metrics_quality,
        "status": status,
        "reviews": reviews_payload,
        "insight": insight_payload,
        "knowledge": knowledge_payload,
        "settings": settings_payload,
    }


@nicegui_app.post("/api/unified/translate")
async def api_unified_translate(request: Request) -> dict:
    body = await request.json()
    text = str((body or {}).get("text") or "").strip()
    force_simplified = bool((body or {}).get("force_simplified"))
    force = bool((body or {}).get("force"))
    if not text:
        return {"ok": False, "error": "empty text", "translated": ""}
    if _is_mostly_chinese(text):
        zh_cache_key = f"zh:{'force' if force_simplified else 'auto'}:{text[:2000]}"
        with _TRANSLATION_CACHE_LOCK:
            zh_cached = None if force else _TRANSLATION_CACHE.get(zh_cache_key)
        if zh_cached:
            return {
                "ok": True,
                "translated": zh_cached,
                "cached": True,
                "detected_language": "zh",
                "normalized_script": "zh-Hans",
            }
        converted = await _to_simplified_chinese(text, force_remote=force_simplified)
        with _TRANSLATION_CACHE_LOCK:
            _TRANSLATION_CACHE[zh_cache_key] = converted
            if len(_TRANSLATION_CACHE) > _TRANSLATION_CACHE_MAX:
                for key in list(_TRANSLATION_CACHE.keys())[: len(_TRANSLATION_CACHE) - _TRANSLATION_CACHE_MAX]:
                    _TRANSLATION_CACHE.pop(key, None)
        return {
            "ok": True,
            "translated": converted,
            "cached": False,
            "detected_language": "zh",
            "normalized_script": "zh-Hans",
        }
    cache_key = text[:2000]
    with _TRANSLATION_CACHE_LOCK:
        cached = None if force else _TRANSLATION_CACHE.get(cache_key)
    if cached:
        return {"ok": True, "translated": cached, "cached": True, "detected_language": "non_zh"}
    try:
        prompt = (
            "请将以下用户评论翻译为简体中文。"
            "只输出译文，不要解释，不要添加前后缀：\n\n"
            f"{text[:4000]}"
        )
        translated = (await chat_with_configured_provider(prompt, max_tokens=800)).strip()
        translated = await _to_simplified_chinese(translated or text)
        if not translated:
            translated = text
        with _TRANSLATION_CACHE_LOCK:
            _TRANSLATION_CACHE[cache_key] = translated
            if len(_TRANSLATION_CACHE) > _TRANSLATION_CACHE_MAX:
                # Keep cache bounded for long-running UI sessions.
                for key in list(_TRANSLATION_CACHE.keys())[: len(_TRANSLATION_CACHE) - _TRANSLATION_CACHE_MAX]:
                    _TRANSLATION_CACHE.pop(key, None)
        return {"ok": True, "translated": translated, "cached": False, "detected_language": "non_zh"}
    except Exception as exc:
        return {"ok": False, "error": str(exc), "translated": text, "fallback": "original_text"}


@nicegui_app.get("/api/unified/settings")
def api_unified_settings() -> dict:
    return {"ok": True, "settings": load_settings(include_secrets=False)}


@nicegui_app.get("/api/unified/fs/dirs")
def api_unified_fs_dirs(path: str = "", limit: int = 300) -> dict:
    current = _resolve_directory_for_listing(path)
    max_items = max(20, min(int(limit or 300), 1200))
    dirs = []
    try:
        children = sorted([item for item in current.iterdir() if item.is_dir()], key=lambda item: item.name.lower())
    except Exception as exc:
        return {"ok": False, "error": str(exc), "current": _display_path(current), "dirs": []}
    for item in children[:max_items]:
        dirs.append(
            {
                "name": item.name,
                "path": str(item),
                "display_path": _display_path(item),
            }
        )
    parent = current.parent if current.parent != current else current
    return {
        "ok": True,
        "current": str(current),
        "current_display": _display_path(current),
        "parent": str(parent),
        "parent_display": _display_path(parent),
        "workspace_root": str(ROOT),
        "workspace_root_display": _display_path(ROOT),
        "count": len(dirs),
        "dirs": dirs,
    }


@nicegui_app.get("/api/unified/knowledge")
def api_unified_knowledge(limit: int = 30) -> dict:
    items = sorted(_load_knowledge_index(), key=lambda item: str(item.get("updated_at") or ""), reverse=True)
    for item in items:
        file_rel = str(item.get("file") or "")
        file_path = ROOT / file_rel if file_rel else Path()
        item["exists"] = bool(file_rel and file_path.exists())
    max_items = max(1, min(int(limit or 30), 300))
    return {"ok": True, "count": len(items), "entries": items[:max_items]}


@nicegui_app.post("/api/unified/knowledge")
async def api_unified_knowledge_add(request: Request) -> dict:
    body = await request.json()
    name = str((body or {}).get("name") or "").strip() or "knowledge_note"
    content = str((body or {}).get("content") or "").strip()
    tags = _as_list((body or {}).get("tags"))
    source_type = str((body or {}).get("source_type") or "manual")
    if not content:
        return {"ok": False, "error": "content is required"}

    try:
        entry = _persist_knowledge_entry(name, content, source_type, tags)
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}
    entries = _load_knowledge_index()
    return {"ok": True, "entry": entry, "entries": sorted(entries, key=lambda item: str(item.get('updated_at') or ''), reverse=True)}


@nicegui_app.post("/api/unified/knowledge/upload")
async def api_unified_knowledge_upload(request: Request) -> dict:
    body = await request.json()
    filename = str((body or {}).get("filename") or "uploaded_knowledge").strip() or "uploaded_knowledge"
    encoded = str((body or {}).get("content_base64") or "")
    inline_content = str((body or {}).get("content") or "")
    if encoded:
        try:
            data = base64.b64decode(encoded, validate=True)
        except Exception as exc:
            return {"ok": False, "error": f"invalid base64: {exc}"}
    else:
        data = inline_content.encode("utf-8")
    if not data:
        return {"ok": False, "error": "empty file"}
    if len(data) > 20 * 1024 * 1024:
        return {"ok": False, "error": "file too large; max 20MB"}
    content = _extract_knowledge_file_text(filename, data).strip()
    if not content:
        return {"ok": False, "error": "no readable text extracted"}
    try:
        entry = _persist_knowledge_entry(filename, content[:500000], "file_upload", tags=["upload", Path(filename).suffix.lower().lstrip(".")])
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}
    entries = _load_knowledge_index()
    return {
        "ok": True,
        "entry": entry,
        "bytes": len(data),
        "chars": len(content),
        "entries": sorted(entries, key=lambda item: str(item.get("updated_at") or ""), reverse=True),
    }


@nicegui_app.post("/api/unified/knowledge/delete")
async def api_unified_knowledge_delete(request: Request) -> dict:
    body = await request.json()
    entry_id = str((body or {}).get("id") or "").strip()
    if not entry_id:
        return {"ok": False, "error": "id is required"}
    entries = _load_knowledge_index()
    matched = next((item for item in entries if str(item.get("id") or "") == entry_id), None)
    if not matched:
        return {"ok": False, "error": "entry not found"}
    file_rel = str(matched.get("file") or "")
    if file_rel:
        file_path = ROOT / file_rel
        try:
            if file_path.exists() and file_path.is_file():
                file_path.unlink()
        except Exception:
            pass
    retained = [item for item in entries if str(item.get("id") or "") != entry_id]
    _save_knowledge_index(retained)
    EVENT_BUS.publish("info", "Knowledge base updated", f"Removed knowledge entry: {matched.get('name') or entry_id}", {"entry_id": entry_id})
    return {"ok": True, "removed_id": entry_id, "entries": sorted(retained, key=lambda item: str(item.get('updated_at') or ''), reverse=True)}


@nicegui_app.post("/api/unified/settings")
async def api_unified_settings_save(request: Request) -> dict:
    body = await request.json()
    if isinstance(body, dict) and body.get("__reset"):
        settings = reset_settings()
    else:
        settings = save_settings(body if isinstance(body, dict) else {})
    processing = settings.get("processing") or {}
    COORDINATOR.configure_limits(
        real_concurrency=int(processing.get("real_concurrency") or 1),
        dry_run_concurrency=int(processing.get("dry_run_concurrency") or 8),
    )
    try:
        sync_interval = int((processing.get("sync_interval_seconds")) or 3600)
    except Exception:
        sync_interval = 3600
    SYNC_MONITOR.set_interval_seconds(sync_interval)
    EVENT_BUS.publish("success", "Settings saved", "Runtime collection settings were updated.")
    return {"ok": True, "settings": settings}


@nicegui_app.get("/api/unified/production-check")
def api_unified_production_check() -> dict:
    settings = load_settings(include_secrets=False)
    status = api_unified_status()
    runs = api_unified_runs(limit=20)
    checks = smoke_check_settings(settings)
    checks["backend"] = {
        "store_count": status.get("store_count", 0),
        "platform_count": len(status.get("platforms", {})),
        "task_count": len(status.get("tasks", [])),
        "run_count": runs.get("count", 0),
        "monitor_running": (status.get("monitor") or {}).get("running", False),
    }
    # Backward-compatible response: keep `production_check`, and expose top-level fields
    # so frontend and external probes can consume a flat schema directly.
    return {
        "ok": checks.get("ok", False),
        "checks": checks.get("checks", []),
        "model_distribution": checks.get("model_distribution", {}),
        "risk_counts": checks.get("risk_counts", {}),
        "backend": checks.get("backend", {}),
        "settings": settings,
        "production_check": checks,
    }


@nicegui_app.post("/api/unified/model-smoke")
async def api_unified_model_smoke(request: Request) -> dict:
    body = await request.json()
    provider = str(body.get("provider") or "") if isinstance(body, dict) else ""
    result = await smoke_test_provider(provider_name=provider or None)
    return {"ok": bool(result.get("ok")), "result": result}


@nicegui_app.post("/api/unified/dry-run")
async def api_unified_dry_run(request: Request) -> dict:
    body = await request.json()
    template = str(body.get("template") or body.get("task_template") or "")
    platform = canonical_platform(str(body.get("platform") or ""))
    task_path = TASK_DIR / template if template else None
    if not task_path or not task_path.exists():
        candidates = [path for path in _task_files() if platform and platform in canonical_platform(path.stem)]
        task_path = candidates[0] if candidates else (_task_files()[0] if _task_files() else None)
    if not task_path or not task_path.exists():
        return {"ok": False, "error": "no task template available"}
    task = load_task(task_path)
    try:
        time_range = validate_week_range(
            start_date=str(body.get("start_date") or ""),
            end_date=str(body.get("end_date") or ""),
            days=int(body["days"]) if body.get("days") is not None else None,
        )
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}
    task = replace(task, time_range=time_range, options={**task.options, "dry_run": True})
    result = COORDINATOR.dry_run(task)
    return {"ok": result.ok, "template": task_path.name, "time_range": time_range.__dict__, "result": result.to_dict()}


@nicegui_app.get("/api/unified/events")
def api_unified_events(since_id: int = 0) -> dict:
    events = EVENT_BUS.list_since(since_id)
    return {"ok": True, "latest_id": EVENT_BUS.latest_id(), "events": events}


@nicegui_app.get("/api/unified/monitor/status")
def api_unified_monitor_status() -> dict:
    return {"ok": True, "monitor": SYNC_MONITOR.status()}


@nicegui_app.post("/api/unified/collect")
async def api_unified_collect(request: Request) -> dict:
    body = await request.json()
    template = str(body.get("template") or "")
    dry_run = bool(body.get("dry_run", True))
    try:
        time_range = validate_week_range(
            start_date=str(body.get("start_date") or ""),
            end_date=str(body.get("end_date") or ""),
            days=int(body["days"]) if body.get("days") is not None else None,
        )
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}
    task_path = TASK_DIR / template if template else None
    if not task_path or not task_path.exists():
        platform = canonical_platform(str(body.get("platform") or ""))
        candidates = [path for path in _task_files() if platform and platform in canonical_platform(path.stem)]
        task_path = candidates[0] if candidates else (_task_files()[0] if _task_files() else None)
    if not task_path or not task_path.exists():
        return {"ok": False, "error": "no task template available"}
    task = load_task(task_path)
    task = replace(task, time_range=time_range, options={**task.options, "dry_run": dry_run})
    if dry_run:
        result = COORDINATOR.dry_run(task)
        EVENT_BUS.publish(
            "success" if result.ok else "error",
            "Manual dry-run completed" if result.ok else "Manual dry-run failed",
            f"{result.platform}: reviews={result.review_count}, stores={result.store_count}",
            {"template": task_path.name, "result": result.to_dict()},
        )
        return {"ok": result.ok, "template": task_path.name, "time_range": time_range.__dict__, "result": result.to_dict()}

    EVENT_BUS.publish(
        "info",
        "Manual collection queued",
        f"{task.platform} / {task.account or task.country}: started in background thread",
        {"template": task_path.name, "time_range": time_range.__dict__},
    )

    def _run_collect_background() -> None:
        result = COORDINATOR.run(task, action="manual_collect")
        EVENT_BUS.publish(
            "success" if result.ok else "error",
            "Manual collection completed" if result.ok else "Manual collection failed",
            f"{result.platform}: reviews={result.review_count}, stores={result.store_count}",
            {"template": task_path.name, "result": result.to_dict()},
        )

    threading.Thread(target=_run_collect_background, daemon=True).start()
    return {
        "ok": True,
        "accepted": True,
        "running": True,
        "template": task_path.name,
        "time_range": time_range.__dict__,
        "message": "manual collection started in background",
    }


@nicegui_app.post("/api/unified/monitor/start")
async def api_unified_monitor_start(request: Request) -> dict:
    body = await request.json()
    templates = body.get("templates") or body.get("task_templates") or [path.name for path in _task_files()]
    templates = [str(item) for item in templates if str(item).endswith(".json")]
    if not templates:
        return {"ok": False, "error": "no task templates provided"}
    try:
        time_range = validate_week_range(
            start_date=str(body.get("start_date") or ""),
            end_date=str(body.get("end_date") or ""),
            days=int(body["days"]) if body.get("days") is not None else 7,
        )
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}
    saved_settings = load_settings(include_secrets=False)
    processing_settings = saved_settings.get("processing") or {}
    COORDINATOR.configure_limits(
        real_concurrency=int(processing_settings.get("real_concurrency") or 1),
        dry_run_concurrency=int(processing_settings.get("dry_run_concurrency") or 8),
    )
    default_interval = int((processing_settings.get("sync_interval_seconds")) or 3600)
    default_workers = int((processing_settings.get("parallel_workers")) or 1)
    state = SYNC_MONITOR.start(
        TASK_DIR,
        templates=templates,
        interval_seconds=int(body.get("interval_seconds", default_interval)),
        dry_run=bool(body.get("dry_run", True)),
        time_range=time_range,
        parallel_workers=int(body.get("parallel_workers", default_workers)),
    )
    return {"ok": True, "monitor": state}


@nicegui_app.post("/api/unified/monitor/run-once")
async def api_unified_monitor_run_once(request: Request) -> dict:
    body = await request.json()
    templates = body.get("templates") or body.get("task_templates") or [path.name for path in _task_files()]
    templates = [str(item) for item in templates if str(item).endswith(".json")]
    if not templates:
        return {"ok": False, "error": "no task templates provided"}
    try:
        time_range = validate_week_range(
            start_date=str(body.get("start_date") or ""),
            end_date=str(body.get("end_date") or ""),
            days=int(body["days"]) if body.get("days") is not None else 7,
        )
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}

    saved_settings = load_settings(include_secrets=False)
    processing_settings = saved_settings.get("processing") or {}
    default_workers = int((processing_settings.get("parallel_workers")) or 1)
    dry_run = bool(body.get("dry_run", False))
    parallel_workers = int(body.get("parallel_workers", default_workers))
    COORDINATOR.configure_limits(
        real_concurrency=int(processing_settings.get("real_concurrency") or 1),
        dry_run_concurrency=int(processing_settings.get("dry_run_concurrency") or 8),
    )
    EVENT_BUS.publish(
        "info",
        "One-shot platform sync queued",
        f"{len(templates)} task template(s), dry_run={dry_run}, workers={parallel_workers}",
        {"templates": templates, "time_range": time_range.__dict__},
    )

    def _run_once_background() -> None:
        try:
            results = SYNC_MONITOR.run_once(
                TASK_DIR,
                templates,
                dry_run=dry_run,
                time_range=time_range,
                parallel_workers=parallel_workers,
            )
            ok_count = sum(1 for item in results if item.get("ok"))
            EVENT_BUS.publish(
                "success" if ok_count == len(results) else "warning",
                "One-shot platform sync completed",
                f"{ok_count}/{len(results)} task template(s) completed successfully",
                {"results": results[-20:]},
            )
        except Exception as exc:
            EVENT_BUS.publish("error", "One-shot platform sync failed", str(exc), {"templates": templates})

    threading.Thread(target=_run_once_background, daemon=True).start()
    return {
        "ok": True,
        "accepted": True,
        "running": True,
        "dry_run": dry_run,
        "templates": templates,
        "time_range": time_range.__dict__,
        "message": "one-shot platform sync started in background",
    }


@nicegui_app.post("/api/unified/monitor/stop")
def api_unified_monitor_stop() -> dict:
    return {"ok": True, "monitor": SYNC_MONITOR.stop()}

# ── 全局样式（黑白配色）─────────────────────────────────────────
ui.add_head_html("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
  * { box-sizing: border-box; }
  body {
    background: #0a0a0a;
    font-family: 'Inter', sans-serif;
    color: #e5e5e5;
  }
  /* 卡片 */
  .nicegui-card {
    background: #141414 !important;
    border: 1px solid #2a2a2a !important;
    border-radius: 10px !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.5) !important;
  }
  /* 输入框 */
  .q-field__control { background: #1a1a1a !important; border-color: #333 !important; }
  .q-field__native, .q-field__input { color: #e5e5e5 !important; }
  .q-field__label { color: #888 !important; }
  /* 选择框下拉 */
  .q-menu { background: #1a1a1a !important; border: 1px solid #333 !important; }
  .q-item { color: #e5e5e5 !important; }
  .q-item:hover { background: #2a2a2a !important; }
  /* 按钮 */
  .q-btn { border-radius: 6px !important; }
  /* 滚动条 */
  ::-webkit-scrollbar { width: 6px; height: 6px; }
  ::-webkit-scrollbar-track { background: #111; }
  ::-webkit-scrollbar-thumb { background: #333; border-radius: 3px; }
  ::-webkit-scrollbar-thumb:hover { background: #555; }
  /* 链接 */
  a { color: #e5e5e5 !important; text-decoration: none !important; }
  a:hover { color: #fff !important; }
  /* 徽章 */
  .q-badge { border-radius: 4px !important; }
  /* 表格 */
  .q-table { background: #141414 !important; color: #e5e5e5 !important; }
  .q-table th { background: #1f1f1f !important; color: #aaa !important; border-bottom: 1px solid #2a2a2a !important; }
  .q-table td { border-bottom: 1px solid #1f1f1f !important; }
  .q-table tr:hover td { background: #1a1a1a !important; }
</style>
""", shared=True)


@ui.page("/")
def index():
    ui.navigate.to("/stitch-static/login_animation/code.html")

@ui.page("/crawler")
def page_crawler():
    ui.navigate.to("/stitch-static/collection_tasks_global/code.html")

@ui.page("/reviews")
def page_reviews():
    ui.navigate.to("/stitch-static/review_workbench_global/code.html")

@ui.page("/reports")
def page_reports():
    ui.navigate.to("/stitch-static/quality_report_global/code.html")

@ui.page("/settings")
def page_settings():
    ui.navigate.to("/stitch-static/safety_audit_global/code.html")

@ui.page("/sentiment")
def page_sentiment():
    ui.navigate.to("/stitch-static/dashboard_global/code.html")

@ui.page("/prototype")
def page_prototype():
    ui.navigate.to("/stitch-static/login_animation/code.html")

@ui.page("/global-ops")
def page_global_ops():
    ui.navigate.to("/stitch-static/collection_tasks_global/code.html")


@ui.page("/legacy")
def page_legacy_dashboard():
    """Legacy NiceGUI shell kept for development only; not the product entry."""
    _render_layout("global_ops")


_NAV_ITEMS = [
    ("📊 仪表盘",   "/",           "dashboard"),
    ("🌐 大一统控制台", "/global-ops", "global_ops"),
    ("🕷️ 爬取管理", "/crawler",    "crawler"),
    ("📝 评论列表", "/reviews",    "reviews"),
    ("📈 统计报表", "/reports",    "reports"),
    ("📡 舆情监控", "/sentiment",  "sentiment"),
    ("🧩 产品原型", "/prototype",  "prototype"),
    ("⚙️ 系统设置", "/settings",   "settings"),
]

def _render_prototype() -> None:
    ui.label("Stitch 产品原型").style("font-size:24px; font-weight:700; color:#fff")
    ui.label("海外门店评论统一采集与客诉洞察平台 · 原始 Stitch 导出页面").style(
        "font-size:13px; color:#888; margin-bottom:12px"
    )
    prototype_pages = [
        ("总览 Dashboard", "/stitch-static/dashboard_global/code.html"),
        ("采集任务", "/stitch-static/collection_tasks_global/code.html"),
        ("门店覆盖", "/stitch-static/store_coverage_global/code.html"),
        ("评论工作台", "/stitch-static/review_workbench_global/code.html"),
        ("平台矩阵", "/stitch-static/platform_matrix_global/code.html"),
        ("质量报告", "/stitch-static/quality_report_global/code.html"),
        ("安全审计", "/stitch-static/safety_audit_global/code.html"),
    ]
    with ui.row().classes("items-center gap-2 mb-3"):
        ui.button("打开完整原型", on_click=lambda: ui.navigate.to("/stitch-static/login_animation/code.html", new_tab=True)).props(
            "flat"
        ).style("background:#fff; color:#000; font-weight:600")
        ui.button("进入可执行控制台", on_click=lambda: ui.navigate.to("/global-ops")).props("flat").style(
            "background:transparent; color:#fff; border:1px solid #444; font-weight:600"
        )
        ui.label("只读原型按钮已接入：导航、导出、任务入口可响应；真实采集在大一统控制台执行").style(
            "font-size:12px; color:#777"
        )
    with ui.row().classes("items-center gap-2 mb-3"):
        for label, href in prototype_pages:
            ui.button(label, on_click=lambda h=href: ui.navigate.to(h, new_tab=True)).props("flat dense").style(
                "background:#171717; color:#bbb; border:1px solid #333; font-size:12px"
            )
    ui.element("iframe").props('src="/stitch-static/dashboard_global/code.html"').classes("w-full").style(
        "height:calc(100vh - 170px); border:1px solid #2a2a2a; "
        "background:#fff; border-radius:8px;"
    )


_PAGE_RENDERERS = {
    "dashboard": dashboard.render,
    "global_ops": global_ops.render,
    "crawler":   crawler.render,
    "reviews":   reviews.render,
    "reports":   reports.render,
    "sentiment": sentiment.render,
    "prototype": _render_prototype,
    "settings":  settings.render,
}


def _render_layout(active_page: str) -> None:
    """渲染整体布局：侧边栏 + 内容区（黑白配色）"""
    with ui.row().classes("w-full min-h-screen").style("background:#0a0a0a"):
        # ── 侧边栏 ────────────────────────────────────────────
        with ui.column().classes("min-h-screen flex flex-col py-6 px-3 gap-1").style(
            "width:220px; background:#111111; border-right:1px solid #222; flex-shrink:0"
        ):
            # Logo
            with ui.row().classes("items-center gap-2 px-2 mb-8"):
                ui.label("◈").style("font-size:20px; color:#fff; font-weight:700")
                with ui.column().classes("gap-0"):
                    ui.label("舆情监控").style("font-size:13px; font-weight:700; color:#fff; line-height:1.2")
                    ui.label("Brand Intelligence").style("font-size:10px; color:#555; line-height:1.2")

            # 分组标签
            ui.label("ANALYTICS").style("font-size:9px; color:#444; font-weight:600; letter-spacing:1.5px; padding:0 8px; margin-top:4px")

            # 导航按钮
            for label, href, page_key in _NAV_ITEMS:
                is_active = page_key == active_page
                icon_map = {
                    "dashboard": "grid_view",
                    "global_ops": "hub",
                    "crawler": "travel_explore",
                    "reviews": "rate_review",
                    "reports": "bar_chart",
                    "sentiment": "radar",
                    "prototype": "dashboard_customize",
                    "settings": "settings",
                }
                icon = icon_map.get(page_key, "circle")

                # 分割线
                if page_key == "settings":
                    ui.separator().style("border-color:#222; margin:8px 0")

                with ui.link(target=href).classes("w-full no-underline"):
                    with ui.row().classes("items-center gap-3 w-full px-2 py-2 rounded-lg").style(
                        ("background:#ffffff; color:#000000;" if is_active
                         else "color:#888888;") +
                        "cursor:pointer; transition:all 0.15s; border-radius:8px;"
                    ):
                        ui.icon(icon).style(
                            f"font-size:16px; {'color:#000' if is_active else 'color:#555'}"
                        )
                        # 只取 emoji 后的中文文字
                        txt = label.split(" ", 1)[-1] if " " in label else label
                        ui.label(txt).style(
                            f"font-size:13px; font-weight:{'600' if is_active else '400'};"
                            f" {'color:#000' if is_active else 'color:#888'}"
                        )

            ui.space()
            # 底部状态
            with ui.column().classes("gap-1 px-2").style("margin-top:auto"):
                with ui.row().classes("items-center gap-2"):
                    ui.label("").style(
                        "width:7px; height:7px; border-radius:50%; background:#4ade80; flex-shrink:0"
                    )
                    ui.label("系统运行中").style("font-size:11px; color:#555")
                ui.label("v2.0.0").style("font-size:10px; color:#333")

        # ── 主内容区 ──────────────────────────────────────────
        with ui.column().classes("flex-1 overflow-auto").style(
            "padding:28px 32px; background:#0a0a0a; min-height:100vh"
        ):
            renderer = _PAGE_RENDERERS.get(active_page, dashboard.render)
            renderer()


if __name__ in {"__main__", "__mp_main__"}:
    ui.run(
        title="Brand Intelligence",
        host="0.0.0.0",
        port=8080,
        reload=False,
        favicon="◈",
        dark=True,
    )
