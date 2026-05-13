"""
Excel 店铺与平台元数据读取模块。

数据源：
1. `门店清单`：基础门店信息、JDE、美团/点评 ID、地址
2. `各门店全渠道周报`：Google Maps / 大众点评 / OpenRice / Keeta 等平台链接或备注
"""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Optional
import logging

import openpyxl
from openpyxl.utils import column_index_from_string

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent
EXCEL_CANDIDATES = (
    PROJECT_ROOT / "港澳&海外客诉汇总底表（3.26-4.1）.xlsx",
    Path(r"C:\Users\Administrator\Desktop\het\港澳&海外客诉汇总底表（3.26-4.1）.xlsx"),
)
SHOP_LIST_SHEET = "门店清单"
WEEKLY_SHEET = "各门店全渠道周报"

PLATFORM_COLUMN_PAIRS = {
    "google_maps": ("K", "L"),
    "dianping": ("M", "N"),
    "openrice": ("O", "P"),
    "keeta": ("Q", "R"),
    "mfood": ("S", "T"),
    "aomi": ("U", "V"),
    "grabfood": ("W", "X"),
    "hungry_panda": ("Y", "Z"),
    "uber_eats": ("AA", "AB"),
    "fantuan": ("AC", "AD"),
    "baedal_minjok": ("AE", "AF"),
}


def resolve_excel_path() -> Path:
    for candidate in EXCEL_CANDIDATES:
        if candidate.exists():
            return candidate
    return EXCEL_CANDIDATES[0]


EXCEL_PATH = resolve_excel_path()


def _clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"None", "无", "/", "-"} else text


def _looks_like_url(value: str) -> bool:
    return value.startswith("http://") or value.startswith("https://")


def _extract_link_and_note(cell) -> tuple[str, str]:
    text = _clean(cell.value)
    hyperlink = cell.hyperlink.target.strip() if cell.hyperlink and cell.hyperlink.target else ""
    if hyperlink:
        note = "" if not text or text == hyperlink or text.startswith("=_XLFN.DISPIMG") else text
        return hyperlink, note
    if _looks_like_url(text):
        return text, ""
    if text.startswith("=_XLFN.DISPIMG"):
        return "", ""
    return "", text


def _copy_rows(rows: tuple[dict, ...]) -> list[dict]:
    return [dict(row) for row in rows]


@lru_cache(maxsize=1)
def read_shop_platform_matrix() -> dict[str, dict]:
    """
    读取 `各门店全渠道周报` 中的平台链接/备注矩阵，按 JDE 建索引。
    """
    workbook_formula = openpyxl.load_workbook(str(EXCEL_PATH), read_only=False, data_only=False)
    workbook_values = openpyxl.load_workbook(str(EXCEL_PATH), read_only=False, data_only=True)
    ws_formula = workbook_formula[WEEKLY_SHEET]
    ws_values = workbook_values[WEEKLY_SHEET]

    matrix: dict[str, dict] = {}
    for row_idx in range(3, ws_values.max_row + 1):
        store_code = _clean(ws_values.cell(row_idx, 4).value)
        if not store_code:
            continue

        shop_name = _clean(ws_values.cell(row_idx, 6).value)
        region = _clean(ws_values.cell(row_idx, 1).value)
        city = _clean(ws_values.cell(row_idx, 3).value)
        period = _clean(ws_values.cell(row_idx, 8).value)
        row_data = {
            "store_code": store_code,
            "weekly_region": region,
            "weekly_city": city,
            "weekly_shop_name": shop_name,
            "weekly_period": period,
        }

        for platform_key, (score_col, link_col) in PLATFORM_COLUMN_PAIRS.items():
            score_value = _clean(ws_values[f"{score_col}{row_idx}"].value)
            link_cell = ws_formula[f"{link_col}{row_idx}"]
            url, note = _extract_link_and_note(link_cell)

            if score_value:
                row_data[f"{platform_key}_rating"] = score_value
            if url:
                row_data[f"{platform_key}_url"] = url
            if note:
                row_data[f"{platform_key}_note"] = note

        matrix[store_code] = row_data

    workbook_formula.close()
    workbook_values.close()
    logger.info("周报平台矩阵已加载：%d 家门店", len(matrix))
    return matrix


@lru_cache(maxsize=1)
def _read_all_shops_cached() -> tuple[dict, ...]:
    workbook = openpyxl.load_workbook(str(EXCEL_PATH), read_only=False, data_only=True)
    worksheet = workbook[SHOP_LIST_SHEET]
    matrix = read_shop_platform_matrix()
    shops: list[dict] = []
    headers = [_clean(worksheet.cell(1, column_idx).value) for column_idx in range(1, worksheet.max_column + 1)]

    for row_idx in range(2, worksheet.max_row + 1):
        values = [_clean(worksheet.cell(row_idx, column_idx).value) for column_idx in range(1, worksheet.max_column + 1)]
        if not any(values):
            continue
        data = dict(zip(headers, values))
        shop_name = data.get("门店", "")
        if not shop_name:
            continue

        store_code = data.get("经营单位代码", "")
        province = data.get("省份", "")
        city = data.get("城市", "")
        merged = {
            "shop_name": shop_name,
            "region": province or city,
            "province": province,
            "city": city,
            "district": data.get("县级市/区", ""),
            "address": data.get("门店地址", ""),
            "meituan_id": data.get("美团外卖ID", ""),
            "meituan_poi_id": data.get("美团ID", ""),
            "dianping_id": data.get("大众点评ID", ""),
            "store_code": store_code,
            "mini_program_code": data.get("小程序编码", ""),
            "business_type": data.get("经营类型", ""),
            "war_zone": data.get("战区", ""),
            "sub_zone": data.get("分区", ""),
            "manager": data.get("区域经理", ""),
            "open_date": data.get("开业日期", ""),
        }
        merged.update(matrix.get(store_code, {}))
        shops.append(merged)

    workbook.close()
    logger.info("共读取 %d 家门店（含平台矩阵）", len(shops))
    return tuple(shops)


def read_shop_list(
    region_filter: Optional[str] = None,
) -> list[dict]:
    """
    读取门店列表，并自动合并周报页中的平台链接与备注。
    """
    shops = _copy_rows(_read_all_shops_cached())
    if not region_filter:
        return shops
    filtered = []
    for shop in shops:
        haystacks = (
            shop.get("region", ""),
            shop.get("province", ""),
            shop.get("city", ""),
            shop.get("sub_zone", ""),
            shop.get("war_zone", ""),
        )
        if any(region_filter in value for value in haystacks):
            filtered.append(shop)
    logger.info("共读取 %d 家%s门店", len(filtered), region_filter)
    return filtered


def get_hk_shop_names() -> list[str]:
    shops = read_shop_list(region_filter="中国香港")
    return [shop["shop_name"] for shop in shops]


def get_all_regions() -> list[str]:
    shops = read_shop_list()
    return list({shop["region"] for shop in shops if shop.get("region")})


if __name__ == "__main__":
    import io
    import sys

    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    print(f"Excel路径: {EXCEL_PATH}")
    shops = read_shop_list()
    print(f"总门店数: {len(shops)}")
    hk = read_shop_list("中国香港")
    print(f"香港门店: {len(hk)}")
    for shop in hk[:5]:
        print(
            f"  {shop['shop_name']} | "
            f"Google={shop.get('google_maps_url','')} | "
            f"点评={shop.get('dianping_url','')} | "
            f"开饭提示={shop.get('openrice_note','')}"
        )
