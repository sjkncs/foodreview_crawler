from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "data" / "store_registry.json"

HEADER_ROW = 2
DATA_START_ROW = 3

PLATFORM_COLUMNS: dict[str, tuple[str, str]] = {
    "google_maps": ("googlemap门店", "Google Maps"),
    "dianping": ("大众用app", "大众点评星级"),
    "openrice": ("开饭啦链接", "开饭星级"),
    "keeta": ("keeta链接", "Keeta"),
    "mfood": ("Mfood链接", "MFoood"),
    "aomi": ("澳觅链接", "澳觅"),
    "grabfood": ("grabfood链接", "Grabfood"),
    "hungry_panda": ("注：不同国家同一账号密码，链接可以直接选取地区", "Hungry Panda"),
    "uber_eats": (
        "https://auth.uber.com/v2/?breeze_init_req_id=a4d1b01d-41ed-4201-a0be-171bd5880aa9&breeze_local_zone=phx5&next_url=https%3A%2F%2Fmerchants.ubereats.com%2Fmanager%2Fpayments%3Fend%3D2024-06-30%26rangeType%3D1%26restaurantUUID%3D107fc9ad-04a7-5075-9be9-76e38fe92758%26start%3D2024-06-17&state=1Ow9aIL23j6V1jgHGivAJeDYhhJOLG0wbJbqUvoik-M%3D",
        "Uber eats",
    ),
    "fantuan": ("饭团链接（选择地区）", "fantuan delivery"),
    "baedal_minjok": ("外卖民族需下载app", "배달의\xa0민족（外卖的民族）"),
}

DEFAULT_ENTRY_URLS = {
    "hungry_panda": "https://merchant-usa.hungrypanda.co/order/appraise",
    "fantuan": "https://merchant.fantuan.ca/#/login",
    "grabfood": "https://merchant.grab.com/portal?source=mrc",
    "keeta": "https://merchant.mykeeta.com/m/web/order#/index",
    "mfood": "https://merchant.o2o.mfoodapp.com/#/appraise/tackout",
    "openrice": "https://www.openrice.com/zh/hongkong/restaurants?chainId=10006678&tabIndex=0",
    "uber_eats": "https://merchants.ubereats.com",
}

HUNGRY_PANDA_ENTRY_BY_COUNTRY = {
    "usa": "https://merchant-usa.hungrypanda.co/order/appraise",
    "us": "https://merchant-usa.hungrypanda.co/order/appraise",
    "canada": "https://merchant-ca.hungrypanda.co/order/appraise",
    "ca": "https://merchant-ca.hungrypanda.co/order/appraise",
    "au": "https://merchant-aus.hungrypanda.co/order/appraise",
    "australia": "https://merchant-aus.hungrypanda.co/order/appraise",
    "uk": "https://merchant-uk.hungrypanda.co/order/appraise",
    "gb": "https://merchant-uk.hungrypanda.co/order/appraise",
    "kr": "https://merchant-kr.hungrypanda.co/order/appraise",
    "korea": "https://merchant-kr.hungrypanda.co/order/appraise",
}

COUNTRY_BY_CODE = {
    "中国澳门": "macau",
    "澳门": "macau",
    "中國澳門": "macau",
    "中国香港": "hong_kong",
    "香港": "hong_kong",
    "加拿大": "canada",
    "马来西亚": "malaysia",
    "美國": "usa",
    "美国": "usa",
    "英国": "uk",
    "英國": "uk",
    "韩国": "kr",
    "韓國": "kr",
    "澳大利亚": "au",
    "澳大利亞": "au",
    "新加坡": "sg",
}

COUNTRY_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("canada", ("加拿大", "Burnaby", "Richmond", "温哥华", "多伦多", "Toronto", "Vancouver")),
    ("malaysia", ("马来西亚", "吉隆坡", "雪兰莪", "槟城", "柔佛", "彭亨", "Malaysia", "Kuala Lumpur")),
    ("usa", ("美国", "纽约", "洛杉矶", "旧金山", "波士顿", "西雅图", "圣地亚哥", "圣何塞", "休斯顿", "湾区", "新泽西", "Virginia", "Irvine")),
    ("uk", ("英国", "伦敦", "曼彻斯特", "爱丁堡", "利兹", "南安普敦", "格拉斯哥", "伯明翰", "利物浦", "London", "Manchester")),
    ("kr", ("韩国", "首尔", "大阪", "Seoul")),
    ("au", ("澳大利亚", "墨尔本", "悉尼", "布里斯班", "Melbourne", "Sydney", "Brisbane")),
    ("hong_kong", ("香港", "Hong Kong")),
    ("macau", ("澳门", "澳門", "Macau")),
    ("sg", ("新加坡", "Singapore", "Vivo", "Jewel", "Orchard")),
]

COUNTRY_DISPLAY = {
    "macau": "澳门",
    "hong_kong": "香港",
    "canada": "加拿大",
    "malaysia": "马来西亚",
    "usa": "美国",
    "uk": "英国",
    "kr": "韩国",
    "au": "澳大利亚",
    "sg": "新加坡",
}

SECRET_TOKENS = ("密码", "账号", "帐号", "password", "passwd", "pwd", "token", "key", "secret", "登录", "總帳號", "总帐号")
ACCOUNT_PLATFORMS = {"hungry_panda", "fantuan", "grabfood", "keeta", "mfood", "uber_eats"}


def clean_text(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return "" if text in {"-", "—", "无", "None", "null"} else text


def is_url(value: str) -> bool:
    return value.startswith(("http://", "https://"))


def is_secret_note(value: str) -> bool:
    lowered = value.lower()
    return any(token.lower() in lowered for token in SECRET_TOKENS)


def country_code(city: str, store_name: str) -> str:
    text = f"{city} {store_name}"
    for raw, code in COUNTRY_BY_CODE.items():
        if raw in text:
            return code
    for code, tokens in COUNTRY_RULES:
        if any(token in text for token in tokens):
            return code
    return clean_text(city) or "unknown"


def account_ref(platform: str, country: str) -> str:
    if platform == "hungry_panda":
        if country == "kr":
            return "hungry_panda:kr"
        if country == "uk":
            return "hungry_panda:uk"
        return "hungry_panda:default"
    if platform == "fantuan":
        country_key = {"canada": "ca", "usa": "us", "au": "au"}.get(country, country)
        return f"fantuan:{country_key}" if country_key in {"ca", "us", "au"} else "fantuan:default"
    if platform == "grabfood":
        return "grabfood:sg" if country == "sg" else "grabfood:my_auro"
    if platform == "keeta":
        return "keeta:default"
    if platform == "mfood":
        return "mfood:default"
    if platform == "uber_eats":
        return "uber_eats:default"
    return f"{platform}:{country or 'default'}"


def default_entry_url(platform: str, country: str) -> str:
    if platform == "hungry_panda":
        key = clean_text(country).lower()
        return HUNGRY_PANDA_ENTRY_BY_COUNTRY.get(key, DEFAULT_ENTRY_URLS["hungry_panda"])
    return DEFAULT_ENTRY_URLS.get(platform, "")


def find_desktop_excel() -> Path:
    desktop = Path.home() / "Desktop"
    preferred = desktop / "海外店鋪.xlsx"
    if preferred.exists():
        return preferred
    matches = [path for path in desktop.glob("*.xlsx") if "海外店" in path.name]
    if matches:
        return matches[0]
    raise FileNotFoundError("海外店铺 Excel not found. Use --excel to specify the workbook path.")


def header_map(row: tuple[Any, ...]) -> dict[str, int]:
    return {clean_text(value): index for index, value in enumerate(row) if clean_text(value)}


def value_by_header(row: tuple[Any, ...], headers: dict[str, int], name: str) -> str:
    index = headers.get(name)
    if index is None or index >= len(row):
        return ""
    return clean_text(row[index])


def platform_entry(row: tuple[Any, ...], headers: dict[str, int], platform: str, country: str) -> dict[str, str]:
    link_header, meta_header = PLATFORM_COLUMNS[platform]
    link = value_by_header(row, headers, link_header)
    meta = value_by_header(row, headers, meta_header)
    fallback_url = default_entry_url(platform, country)
    entry: dict[str, str] = {}
    if is_url(link):
        entry["url"] = link
    elif link and not is_secret_note(link):
        entry["note"] = link
    elif link and fallback_url:
        entry["url"] = fallback_url
    if meta and not is_secret_note(meta):
        entry["meta"] = meta
    if platform in ACCOUNT_PLATFORMS and entry and not entry.get("url") and fallback_url:
        entry["url"] = fallback_url
    if platform in ACCOUNT_PLATFORMS and (entry or link or meta):
        entry["account_ref"] = account_ref(platform, country)
    return entry


def build_registry(excel: Path, sheet: str = "") -> dict[str, Any]:
    wb = load_workbook(excel, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb["各门店全渠道周报"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < DATA_START_ROW:
        raise ValueError("Excel has no store rows.")
    headers = header_map(rows[HEADER_ROW - 1])
    stores: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[DATA_START_ROW - 1 :], start=DATA_START_ROW):
        jde = value_by_header(row, headers, "JDE")
        store_name = value_by_header(row, headers, "门店名称")
        if not jde or not store_name:
            continue
        city = value_by_header(row, headers, "城市")
        code = country_code(city, store_name)
        platforms = {
            platform: entry
            for platform in PLATFORM_COLUMNS
            if (entry := platform_entry(row, headers, platform, code))
        }
        stores.append(
            {
                "jde": jde,
                "store_name": store_name,
                "country": COUNTRY_DISPLAY.get(code, code),
                "country_code": code,
                "city": city,
                "region": value_by_header(row, headers, "区域"),
                "province": value_by_header(row, headers, "省份"),
                "supervisor": value_by_header(row, headers, "督导"),
                "opened_at": value_by_header(row, headers, "开业日期"),
                "source_row": row_index,
                "platforms": platforms,
            }
        )
    return {
        "schema_version": 2,
        "source_excel": str(excel),
        "source_sheet": ws.title,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "store_count": len(stores),
        "stores": stores,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build unified HEYTEA overseas store registry from Excel.")
    parser.add_argument("--excel", default="", help="海外店铺 Excel path.")
    parser.add_argument("--sheet", default="", help="Sheet name; defaults to 各门店全渠道周报.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output registry JSON path.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    excel = Path(args.excel) if args.excel else find_desktop_excel()
    registry = build_registry(excel, args.sheet)
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(registry, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "output": str(output), "store_count": registry["store_count"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
